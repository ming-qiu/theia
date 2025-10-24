#!/usr/bin/env python3
"""
Video File Renamer
Renames video files based on shot list from Excel file.
"""

import re
import argparse
from pathlib import Path
from openpyxl import load_workbook


def parse_filename(filename):
    """
    Extract cut order number from filename like V1-0001_.Editorial.Reference_20251019
    Returns the cut order number (e.g., '0001') or None if pattern doesn't match.
    """
    pattern = r'V\d+-(\d{4})_'
    match = re.match(pattern, filename)
    if match:
        return match.group(1)
    return None


def load_shot_list(excel_path, sheet_name='Shots'):
    """
    Load the shot list from Excel file using openpyxl.
    Returns a dictionary mapping cut order (as int) to shot code.
    """
    try:
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        
        if sheet_name not in workbook.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in workbook")
            print(f"Available sheets: {', '.join(workbook.sheetnames)}")
            return None
        
        sheet = workbook[sheet_name]
        
        # Create mapping dictionary
        # Column B = index 2, Column D = index 4
        shot_mapping = {}
        
        # Skip header row, start from row 2
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 4:
                continue
            
            cut_order = row[1]  # Column B (0-indexed: column 1)
            shot_code = row[3]  # Column D (0-indexed: column 3)
            
            # Skip rows with missing data
            if cut_order is not None and shot_code is not None:
                try:
                    # Convert cut order to int for matching
                    cut_order_int = int(cut_order)
                    shot_mapping[cut_order_int] = str(shot_code).strip()
                except (ValueError, TypeError):
                    continue
        
        workbook.close()
        return shot_mapping
    
    except FileNotFoundError:
        print(f"Error: Excel file not found: {excel_path}")
        return None
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return None


def generate_new_filename(original_filename, cut_order_str, shot_code):
    """
    Generate new filename by replacing Vx-xxxx_ with {cut_order}_{shot_code}
    """
    # Extract the part after the original prefix
    pattern = r'V\d+-\d{4}_(.*)'
    match = re.match(pattern, original_filename)
    
    if match:
        remainder = match.group(1)
        # Format: 3-digit cut order + underscore + shot code + underscore + remainder
        new_filename = f"{int(cut_order_str):03d}_{shot_code}{remainder}"
        return new_filename
    
    return None


def rename_videos(directory, excel_path, sheet_name='Shots', dry_run=False):
    """
    Main function to rename video files based on shot list.
    """
    # Load shot list from Excel
    print(f"Loading shot list from: {excel_path}")
    shot_mapping = load_shot_list(excel_path, sheet_name)
    
    if shot_mapping is None:
        print("Failed to load shot list. Exiting.")
        return
    
    print(f"Loaded {len(shot_mapping)} shots from Excel file\n")
    
    # Find all video files in directory
    video_dir = Path(directory)
    if not video_dir.exists():
        print(f"Directory not found: {directory}")
        return
    
    # Common video extensions
    video_extensions = {'.mp4', '.mov', '.avi', '.mxf', '.mkv', '.m4v'}
    video_files = [f for f in video_dir.iterdir() 
                   if f.is_file() and f.suffix.lower() in video_extensions]
    
    print(f"Found {len(video_files)} video files\n")
    
    renamed_count = 0
    skipped_count = 0
    
    for video_file in sorted(video_files):
        filename = video_file.name
        
        # Parse cut order from filename
        cut_order_str = parse_filename(filename)
        
        if cut_order_str is None:
            print(f"SKIP: {filename} (doesn't match expected pattern)")
            skipped_count += 1
            continue
        
        # Look up shot code
        cut_order_int = int(cut_order_str)
        shot_code = shot_mapping.get(cut_order_int)
        
        if shot_code is None:
            print(f"SKIP: {filename} (cut order {cut_order_int} not found in shot list)")
            skipped_count += 1
            continue
        
        # Generate new filename
        new_filename = generate_new_filename(filename, cut_order_str, shot_code)
        
        if new_filename is None:
            print(f"SKIP: {filename} (failed to generate new name)")
            skipped_count += 1
            continue
        
        new_path = video_file.parent / new_filename
        
        if dry_run:
            print(f"WOULD RENAME: {filename}")
            print(f"          TO: {new_filename}\n")
        else:
            try:
                video_file.rename(new_path)
                print(f"RENAMED: {filename}")
                print(f"     TO: {new_filename}\n")
                renamed_count += 1
            except Exception as e:
                print(f"ERROR renaming {filename}: {e}\n")
                skipped_count += 1
    
    # Summary
    print("=" * 60)
    if dry_run:
        print("DRY RUN COMPLETE - No files were actually renamed")
    else:
        print(f"COMPLETE - {renamed_count} files renamed")
    print(f"Skipped: {skipped_count} files")


def main():
    parser = argparse.ArgumentParser(
        description='Rename video files based on shot list from Excel'
    )
    parser.add_argument(
        'directory',
        help='Directory containing video files'
    )
    parser.add_argument(
        '--excel',
        default='shot_list.xlsx',
        help='Path to Excel file (default: shot_list.xlsx)'
    )
    parser.add_argument(
        '--sheet',
        default='Shots',
        help='Sheet name in Excel file (default: Shots)'
    )
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Print proposed changes without actually renaming files'
    )
    
    args = parser.parse_args()
    
    rename_videos(
        args.directory,
        args.excel,
        args.sheet,
        dry_run=args.dry_run
    )


if __name__ == '__main__':
    main()