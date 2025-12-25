"""
DaVinci Resolve - Import VFX Shot Codes from Excel to Timeline Subtitles
Reads Excel file and creates SRT subtitle files based on selected columns.
Can also add frame counter videos to timeline shots.
"""

import sys, os
import argparse
from timecode import Timecode
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

try:
    import DaVinciResolveScript as dvr
except ImportError:
    dvr = None


def parse_args():
    parser = argparse.ArgumentParser(description='Create SRT subtitle files for shot metadata and attach frame counters to timeline')
    parser.add_argument('--sheet', default = "clip_inventory.xlsx", help = "Excel sheet that contains metadata")
    parser.add_argument('--metadata-from', default="G", help='Column letter of the first metadata column (default: G)')
    parser.add_argument('--metadata-to', default='G', help='Column letter of the last metadata column (default: G)')
    parser.add_argument('--frame-counter', type=str, help='Path to frame counter video file')
    parser.add_argument('--first-frame', type=int, help='Starting frame number for frame counters')
    parser.add_argument('--fps', type=float, default=24.0, help='Timeline FPS (default: 24)')

    return parser.parse_args()

def add_frame_counters(sheet_path, frame_counter_path, first_frame, fps):
    """Add frame counter videos to timeline based on shot timings from Excel."""
    
    if not dvr:
        print("ERROR: DaVinci Resolve API not available")
        sys.exit(1)
    
    # Connect to Resolve
    resolve = dvr.scriptapp("Resolve")
    project = resolve.GetProjectManager().GetCurrentProject()
    timeline = project.GetCurrentTimeline()
    mediapool = project.GetMediaPool()
    
    print(f"Timeline: {timeline.GetName()}")
    
    # Import frame counter to media pool
    abs_frame_counter_path = os.path.abspath(frame_counter_path)
    print(f"Importing frame counter: {abs_frame_counter_path}")
    imported = mediapool.ImportMedia([abs_frame_counter_path])
    if not imported:
        print("ERROR: Failed to import frame counter video")
        sys.exit(1)
    
    frame_counter_item = imported[0]
    fc_first_frame = Timecode(fps, frame_counter_item.GetClipProperty("Start TC")).frames - 1

    # Get number of video tracks and create new track
    num_tracks = timeline.GetTrackCount("video")
    target_track = num_tracks + 1
    print(f"Adding frame counters to track {target_track}")
    timeline.AddTrack("video", None)
    
    # Load Excel to get shot timings
    wb = load_workbook(sheet_path)
    ws = wb.active
    
    # Read shot data
    clips_to_add = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 5:
            continue
        
        record_tc_in = Timecode(fps, str(row[3]))   # Column D
        record_tc_out = Timecode(fps, str(row[4]))  # Column E
        
        shot_duration = record_tc_out.frames - record_tc_in.frames
        
        clips_to_add.append({
            "mediaPoolItem": frame_counter_item,
            "startFrame": first_frame - fc_first_frame,
            "endFrame": first_frame - fc_first_frame + shot_duration,
            "trackIndex": target_track,
            "recordFrame": record_tc_in.frames - 1
        })
    
    print(f"Adding {len(clips_to_add)} frame counter clips...")
    result = mediapool.AppendToTimeline(clips_to_add)
    
    if result:
        print(f"Success! Added {len(result)} frame counter clips to track {target_track}")
    else:
        print("ERROR: Failed to add clips to timeline")

def create_srt_file(ws, output_path, column_index, fps):
    """Create SRT file from specific Excel column."""
    
    # column_index = column_index_from_string(column_letter) - 1

    # Read data from Excel
    subtitles = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= column_index:
            continue
        
        content = row[column_index]
        if content and str(content).strip():
            subtitles.append({
                'tc_in': Timecode(fps, str(row[3])),    # Column D
                'tc_out': Timecode(fps, str(row[4])),   # Column E
                'text': str(content).strip()
            })
    
    if not subtitles:
        return None
    
    # Create SRT content
    srt_lines = []
    for idx, sub in enumerate(subtitles, start=1):
        # Convert to SRT format: HH:MM:SS,mmm
        def to_srt(tc):
            total_sec = tc.frames / float(tc.framerate)
            h = int(total_sec // 3600)
            m = int((total_sec % 3600) // 60)
            s = int(total_sec % 60)
            ms = int((total_sec % 1) * 1000)
            return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"
        
        srt_lines.append(str(idx))
        srt_lines.append(f"{to_srt(sub['tc_in'])} --> {to_srt(sub['tc_out'])}")
        srt_lines.append(sub['text'])
        srt_lines.append("")
    
    # Write SRT file
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(srt_lines))
    
    return output_path

def main():
    
    args = parse_args()
    
    # Handle frame counter mode
    if args.frame_counter:
        if not args.first_frame:
            print("ERROR: --first-frame required when using --frame-counter")
            sys.exit(1)
        add_frame_counters(args.sheet, args.frame_counter, args.first_frame, args.fps)
    
    # Check at least one flag is set for SRT mode
    if not (args.metadata_from or args.metadata_to):
        print("ERROR: At least one metadata column must be specified\n")
        print("Usage examples:")
        print(f"  python {sys.argv[0]} --metadata-from G --metadata-to J")
        sys.exit(1)

    # Load sheet once
    wb = load_workbook(args.sheet)
    ws = wb.active
    
    # Convert column letters to indices
    col_from_idx = column_index_from_string(args.metadata_from.upper())
    col_to_idx = column_index_from_string(args.metadata_to.upper())
    
    # Create SRT files for each column
    for col_idx in range(col_from_idx, col_to_idx + 1):
        # Get column header from first row
        column_header = ws.cell(row=1, column=col_idx).value
        if not column_header:
            continue
        
        # Create output filename from header
        output_path = f"{column_header}.srt"
        if create_srt_file(ws, output_path, col_idx - 1, args.fps):
            print(f"  Created: {output_path}")

if __name__ == "__main__":
    main()