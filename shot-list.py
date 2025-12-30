import os
import sys
import math
import argparse
import json
from collections import defaultdict, namedtuple

# Resolve API
try:
    import DaVinciResolveScript as dvr
except Exception as e:
    print("ERROR: Could not import DaVinciResolveScript. Run from Resolve's Python.", file=sys.stderr)
    raise

# Third-party libs
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from timecode import Timecode
except ImportError as e:
    print("ERROR: Missing dependency. Install with: pip install openpyxl timecode", file=sys.stderr)
    raise

# -------- Utils --------

def parse_args():
    p = argparse.ArgumentParser(description="Export VFX shots/elements from a Resolve timeline to Excel.")
    p.add_argument("--timeline", default = None, help="Timeline name of the current edit (default: current active timeline)")
    p.add_argument("--output", default = "shot_list.xlsx", help="Output Excel path (default: shot_list.xlsx)")
    p.add_argument("--bottom", default = 1, help="Track number of the bottom video layer (default: 1)")
    p.add_argument("--top", default = 4, help="Track number of the top video layer (default: 1)")
    p.add_argument("--counter-track", default = 5, help="Track number of the frame counter (default: 5)")
    p.add_argument("--work-handle", type = int, default=8, help="WORK_HANDLE frames (default: 8)")
    p.add_argument("--scan-handle", type = int, default=24, help="SCAN_HANDLE frames (default: 24)")
    #p.add_argument("--half-frame", action='store_true', help="Apply half-frame offset correction (default False)")
    p.add_argument("--old-timeline", default = None, help="Timeline name of the last edit (default: None)")

    return p.parse_args()

def resolve_app():
    app = dvr.scriptapp("Resolve")
    if not app:
        raise RuntimeError("Could not acquire Resolve app. Run inside Resolve.")
    return app

def get_project(app):
    pm = app.GetProjectManager()
    proj = pm.GetCurrentProject()
    if not proj:
        raise RuntimeError("No Resolve project open.")
    return proj

def get_timeline(proj, name=None):
    if name:
        for i in range(1, proj.GetTimelineCount() + 1):
            tl = proj.GetTimelineByIndex(i)
            if tl and tl.GetName() == name:
                return tl
        raise RuntimeError(f"Timeline '{name}' not found.")
    tl = proj.GetCurrentTimeline()
    if not tl:
        raise RuntimeError("No active timeline.")
    return tl

def get_timeline_fps(timeline):
    # Try timeline setting; fall back to 24
    fps_str = timeline.GetSetting("timelineFrameRate") or ""
    try:
        # preserve decimals for 23.976 etc.
        fps = float(fps_str)
        # timecode lib accepts strings; but we'll keep float and format back when constructing
        return fps
    except Exception:
        return 24.0

def fps_to_str(fps):
    # timecode accepts strings like "23.976", "24", "25", "29.97"
    # Keep up to 3 decimals (common broadcast values)
    if float(fps).is_integer():
        return str(int(round(fps)))
    return f"{fps:.3f}".rstrip('0').rstrip('.')

def _fmt_fps(fps):
    # pretty print fps (int if whole, else 2 decimals)
    if float(fps).is_integer():
        return str(int(round(fps)))
    return f"{fps:.2f}".rstrip('0').rstrip('.')

def safe_get(d, k, default=None):
    try:
        return d.get(k, default)
    except Exception:
        return default

def subtitle_text(item):
    # Prefer subtitle "Text" if exposed, else name
    props = item.GetProperty() or {}
    text = safe_get(props, "Text", item.GetName() or "").strip()
    
    # Extract just the shot code (first non-whitespace token)
    # Split by newline, space, or tab and take the first part
    if text:
        # Split by any whitespace (space, tab, newline, etc.)
        parts = text.split()
        if parts:
            return parts[0]
    
    return text

def element_name_for_track(idx):
    if idx == 1:
        return "ScanBg"
    return f"ScanFg{idx-1:02d}"  # 2->ScanFg01, 3->ScanFg02, ...

def first_nonempty_subtitle_track(timeline):
    n = timeline.GetTrackCount("subtitle")
    all_items = []
    chosen_index = None
    for i in range(1, n + 1):
        items = timeline.GetItemListInTrack("subtitle", i) or []
        if items:
            all_items = items
            chosen_index = i
            break
    return chosen_index, all_items

def get_reel_name(clip_props):
    for k in ("Clip Name", "File Name", "Reel Name"):
        v = clip_props.get(k)
        if v:
            return v
    return "placeholder" # Ming: Resolve API doesn't expose these for offline clips // Oct 18, 2025

def get_sequence_name(shot_code):
    if shot_code.find('_') != -1:
        return shot_code.split('_')[0]
    elif shot_code.find('-') != -1:
        return shot_code.split('-')[0]
    else:
        return 'sequence_name'

def shot_editorial_name_from_bg(bg_elements):
    """Pick the ScanBg reel name for the shot (earliest bg element)."""
    if not bg_elements:
        return ""
    e = sorted(bg_elements, key=lambda x: (x["TimelineStart"], x["TimelineEnd"]))[0]
    return e.get("ReelName", "") or ""

def best_bg_cut_in_tc(bg_elements, fps):
    """Choose Cut In TC from ScanBg (track 1). Prefer the element that starts closest to the shot start."""
    if not bg_elements:
        return ""
    # They are already filtered to the shot; pick earliest ClipInFrames
    e = sorted(bg_elements, key=lambda x: x["ClipInFrames"])[0]
    return e["ClipInTC"]

def get_clip_tc(timeline_item, fps):
    """
    Works for online/offline items. Assumes clip fps == timeline fps.
    Returns: dict with ClipInTC, ClipInFrames, ClipOutTC, ClipOutFrames
    """

    tin = timeline_item.GetSourceStartTime()
    tout = timeline_item.GetSourceEndTime()
    # print (tin, tout)

    if tin is not None and tout is not None:
                # Calculate actual source duration from source times
        source_dur_seconds = tout - tin
        #print ("source_dur_seconds =", source_dur_seconds)
        source_dur_frames = source_dur_seconds * fps
        #print ("source_dur_frames =", source_dur_frames)

        # Get timeline duration (how long it plays on timeline)
        timeline_dur = timeline_item.GetDuration(False) or 0
        #print ("timeline_dur =", timeline_dur)

        # Calculate retime speed: source_frames / timeline_frames
        speed = (source_dur_frames - 1) / (timeline_dur - 1) if timeline_dur > 0 else 1.0

        if abs(speed - 1) < 1e-3:
            speed = 1

        # The actual source frames consumed
        actual_source_frames = timeline_dur * speed
        
        src_in_frames = int(round(tin * fps))
        src_out_frames = int(round(src_in_frames + actual_source_frames))
        
    elif(timeline_item.GetDuration(False)):
        dur = timeline_item.GetDuration(False)
        src_in_frames = 0
        src_out_frames = int(dur)
    else:
        src_in_frames = 0
        src_out_frames = 1

    return {
        "ClipInTC":        repr(Timecode(fps_to_str(fps), frames = src_in_frames)),
        "ClipInFrames":  src_in_frames,
        "ClipOutTC":       repr(Timecode(fps_to_str(fps), frames = src_out_frames)),
        "ClipOutFrames": src_out_frames,
    }

def _is_back_to_back(prev_src_out, curr_src_in, tol=1):
    # Treat exact or +1 frame as “back to back”
    return curr_src_in == prev_src_out or curr_src_in == prev_src_out + 1 or abs(curr_src_in - prev_src_out) <= tol

def _fmt_percent(val):
    f = float(str(val).strip())
    p = f * 100.0
    # show integers when clean, otherwise a compact decimal
    return f"{int(round(p))}%" if abs(p - round(p)) < 1e-6 else f"{p:.2f}%"

def retime_summary(elements_by_track, fps):
    """
    For each track's elements (within a shot), detect non-linear retimes:
    - Group consecutive clips by same ReelName when their source frames are back-to-back
    - Merge grouped clips into a single element representing the non-linear retime
    - For merged elements, calculate frame mappings showing timeline_frame -> source_frame
    - Write RetimeSummary as: "x1 -> y1, x2 -> y2, ..." where x is timeline frame and y is source frame
    - Sets HasRetime (True if any segment speed != 1 or non-linear retime detected)
    - MODIFIES elements_by_track in place, replacing groups with merged elements
    """
    for track_num, track in elements_by_track.items():
        # Sort by timeline order
        track.sort(key=lambda e: (e["TimelineStart"], e["TimelineEnd"]))

        # First pass: compute per-clip source/timeline info and speed
        for clip in track:
            ti = clip["TimelineItem"]
            # Use already-calculated source frames from element dict
            src_frames = max(0, int(clip["ClipOutFrames"] - clip["ClipInFrames"]))
            
            # Timeline duration (frames)
            tl_frames = int(ti.GetDuration(False) or (clip["TimelineEnd"] - clip["TimelineStart"]))
            
            # Calculate speed
            speed = (src_frames / tl_frames) if tl_frames > 0 else None
            
            clip["SourceDur"] = src_frames
            clip["TimelineDur"] = tl_frames
            clip["Speed"] = speed
            clip["RetimeFPS"] = (fps * speed) if speed is not None else None
            clip["RetimeSummary"] = ""
            clip["HasRetime"] = (speed is not None) and (abs(speed - 1.0) > 1e-3)

        # Second pass: group by reel & back-to-back source ranges, then merge
        merged_track = []
        i = 0
        n = len(track)
        
        while i < n:
            group = [track[i]]
            reel = track[i].get("ReelName", "")
            j = i + 1
            
            # Build group of consecutive clips with same reel and back-to-back source
            while j < n:
                same_reel = (track[j].get("ReelName", "") == reel)
                if not same_reel:
                    break
                if not _is_back_to_back(group[-1]["ClipOutFrames"], track[j]["ClipInFrames"]):
                    break
                group.append(track[j])
                j += 1

            # Check if this group has any retime
            any_retime = any(g["HasRetime"] for g in group)
            
            if any_retime and len(group) > 1:
                # Non-linear retime detected - merge into single element
                first = group[0]
                last = group[-1]
                
                # Build frame mapping
                mappings = []
                current_timeline_frame = first["ClipIn"]
                current_source_frame = first["ClipIn"]
                
                # Add first mapping point
                mappings.append(f"{current_timeline_frame} -> {current_source_frame}")
                current_source_frame -= 1
                current_timeline_frame -= 1
                
                # Process each clip in the group to build mappings
                for g in group:
                    timeline_end = current_timeline_frame + g["TimelineDur"]
                    source_end = current_source_frame + g["SourceDur"]
                    mappings.append(f"{timeline_end} -> {source_end}")
                    current_timeline_frame = timeline_end
                    current_source_frame = source_end
                
                summary = ", ".join(mappings)
                
                # Create merged element based on first clip, spanning the entire group
                merged_element = {
                    "TrackIndex": first["TrackIndex"],
                    "ShotCode": first["ShotCode"],
                    "ElementName": first["ElementName"],
                    "TimelineItem": first["TimelineItem"],  # Keep first clip's timeline item
                    "TimelineStart": first["TimelineStart"],
                    "TimelineEnd": last["TimelineEnd"],
                    "ClipIn": first["ClipIn"],
                    "ClipOut": last["ClipOut"],
                    "ClipInTC": first["ClipInTC"],
                    "ClipOutTC": last["ClipOutTC"],
                    "ClipInFrames": first["ClipInFrames"],
                    "ClipOutFrames": last["ClipOutFrames"],
                    "RetimeSummary": summary,
                    "ScaleRepo": first["ScaleRepo"],
                    "HasRetime": True,
                    "ReelName": reel,
                    "Props": first["Props"],
                    "ClipProps": first["ClipProps"],
                    "HeadIn": first["HeadIn"],
                    "TailOut": last["TailOut"],
                    "SourceDur": sum(g["SourceDur"] for g in group) - 1,
                    "TimelineDur": sum(g["TimelineDur"] for g in group) - 1,
                    "Speed": None,  # Non-linear, so no single speed value
                    "RetimeFPS": None,  # Non-linear, so no single fps value
                }
                
                merged_track.append(merged_element)
                    
            elif any_retime:
                # Single clip with retime - keep as is with simple format
                for g in group:
                    if g["HasRetime"]:
                        g["RetimeSummary"] = _fmt_percent(g['Speed'])
                    merged_track.append(g)
            else:
                # No retime - keep all clips as separate elements
                merged_track.extend(group)
            
            i = j
        
        # Replace the track with merged version
        elements_by_track[track_num] = merged_track
    
    print('\n')

def summarize_scale_repo(props):
    zx = safe_get(props, "ZoomX")
    px = safe_get(props, "PositionX")
    py = safe_get(props, "PositionY")
    parts = []
    if zx != 1:
        parts.append(f"Scale: {_fmt_percent(zx)}")
    if px is not None or py is not None:
        parts.append(f"Repo: {px},{py}")
    return " ".join(parts)

def read_edit(args, app, project, timeline, fps):
    # Subtitle items (VFX shots)
    sub_idx, sub_items = first_nonempty_subtitle_track(timeline)
    if not sub_items:
        raise RuntimeError("No subtitle items found. Place shot codes on a subtitle track.")

    # Sort subtitles by start frame; build Cut Order as 1..N
    sub_items_sorted = sorted(sub_items, key=lambda clip: clip.GetStart(False))

    # Video tracks and element naming
    v_tracks = timeline.GetTrackCount("video")
    element_labels = {i: element_name_for_track(i) for i in range(1, v_tracks + 1)}

    shots_rows = []
    elements_rows = []

    # For video tracks {bottom ~ top}, pre-pull all video items per track for faster filtering
    track_items = {}
    for i in range(int(args.bottom), int(args.top) + 1):
        track_name = (timeline.GetTrackName("video", i) or "").lower()
        track_items[i] = timeline.GetItemListInTrack("video", i) or []
    counters = timeline.GetItemListInTrack("video", int(args.counter_track))

    cut_order = 0

    for sub in sub_items_sorted:
        shot_code = subtitle_text(sub)
        if not shot_code:
            # Skip empty subtitle items
            continue

        cut_order += 1
        print (f"==== Cut {cut_order} ====")

        sub_start = sub.GetStart(False)
        sub_end   = sub.GetEnd(False)
        cut_dur = sub_end - sub_start

        
        for counter in counters:
            counter_start = counter.GetStart(False)
            counter_end = counter.GetEnd(False)
            if counter_start >= sub_start and counter_end <= sub_end:
                tc_info = get_clip_tc(counter, fps)
                print (tc_info)
                cut_in  = tc_info['ClipInFrames']
                cut_out  = tc_info['ClipOutFrames'] - 1

        # TO DO: if counter is not available, read BG clips, calculate cut_in with BG in TC + shot_list.xlsx from the last edit

        # Collect elements contained fully inside [sub_start, sub_end]
        elements_by_track = defaultdict(list)

        for track in range(int(args.bottom), int(args.top) + 1):
            if track == args.counter_track:
                continue
            for clip in (track_items.get(track) or []):
                clip_start = clip.GetStart(False)
                clip_end = clip.GetEnd(False)
                if clip_start >= sub_start and clip_end <= sub_end:
                    # Clip & media props
                    mpi = clip.GetMediaPoolItem()
                    clip_props = (mpi.GetClipProperty() if mpi else {}) or {}

                    reel = clip.GetName()

                    # Clip In/Out TC/frames from clip properties (fps == timeline fps)
                    tc_info = get_clip_tc(clip, fps)
                    print (tc_info)

                    clip_in_tc         = tc_info["ClipInTC"]
                    clip_out_tc        = tc_info["ClipOutTC"]
                    clip_in_tc_frames  = tc_info["ClipInFrames"]
                    clip_out_tc_frames = tc_info["ClipOutFrames"]


                    # Clip In/Out in the VFX "Cut" system (pre-SG shift)
                    clip_in  = int(cut_in + (clip_start - sub_start))
                    clip_out = int(clip_in + (clip_out_tc_frames  - clip_in_tc_frames)) - 1

                    # Summaries (from timeline item props)
                    props = clip.GetProperty() or {}
                    scalerpo_sum = summarize_scale_repo(props)

                    # inside your loop where you append each element dict
                    elements_by_track[track].append({
                        "TrackIndex": track,
                        "ShotCode": shot_code,
                        "ElementName": element_labels[track],
                        "TimelineItem": clip,
                        "TimelineStart": clip_start,
                        "TimelineEnd": clip_end,
                        "ClipIn": clip_in,
                        "ClipOut": clip_out,
                        "ClipInTC": clip_in_tc,
                        "ClipOutTC": clip_out_tc,
                        "ClipInFrames": clip_in_tc_frames,
                        "ClipOutFrames": clip_out_tc_frames,
                        "ClipDuration": clip_out - clip_in + 1, 
                        "HasRetime": False,
                        "RetimeSummary": "",
                        "ScaleRepo": scalerpo_sum,
                        "ReelName": reel,
                        "Props": props,
                        "ClipProps": clip_props,
                        "HeadIn": int(clip_in - args.scan_handle),
                        "TailOut": int((clip_out_tc_frames - clip_in_tc_frames) + clip_in + args.scan_handle),
                    })

        # Shot Cut In TC comes from ScanBg (track 1) element's Clip In TC
        bg_elems = elements_by_track.get(1, [])
        cut_in_tc = best_bg_cut_in_tc(bg_elems, fps) if bg_elems else ""     

        # Editorial Name of the shot
        shot_editorial_name = shot_editorial_name_from_bg(bg_elems)

        # Work handles
        work_in  = int(cut_in  - args.work_handle)
        work_out = int(cut_out + args.work_handle)

        retime_summary(elements_by_track, fps)

        # Flags for Shots sheet
        bg_retime = "x" if any(e["HasRetime"] for e in elements_by_track.get(1, [])) else ""
        fg_retime = "x" if any(
            e["HasRetime"]
            for track, lst in elements_by_track.items()
            if track >= 2 for e in lst
        ) else ""

        # Shots row
        shots_rows.append({
            "Sequence": get_sequence_name(shot_code),
            "CutOrder": cut_order,
            "EditorialName": shot_editorial_name,
            "ShotCode": shot_code,
            "ChangeToCut": None,
            "WorkIn": work_in,
            "CutIn": int(cut_in),
            "CutOut": int(cut_out),
            "WorkOut": work_out,
            "CutDuration": cut_dur,
            "BgRetime": bg_retime,
            "FgRetime": fg_retime,
            "CutInTC": cut_in_tc,
        })

        # Elements rows (ordered by track: 1..N)
        for track in range(1, v_tracks + 1):
            for e in sorted(elements_by_track.get(track, []), key=lambda x: (x["TimelineStart"], x["TimelineEnd"])):
                elements_rows.append({
                    "Sequence": get_sequence_name(shot_code),
                    "CutOrder": cut_order,
                    "EditorialName": e["ReelName"],
                    "ShotCode": shot_code,
                    "Element": e["ElementName"],
                    "ShotCutIn": int(cut_in),
                    "ShotCutOut": int(cut_out),
                    "ClipInTC": e["ClipInTC"],
                    "ClipInFrames": int(e["ClipInFrames"]),
                    "ClipIn": int(e["ClipIn"]),
                    "ClipOut": int(e["ClipOut"]),
                    "ClipOutFrames": int(e["ClipOutFrames"]),
                    "ClipOutTC": e["ClipOutTC"],
                    "ClipDuration": e['ClipDuration'],
                    "Retime": e["RetimeSummary"],
                    "ScaleRepo": e["ScaleRepo"],
                })

    # Sort shots/elements by Cut Order (already built that way, but enforce)
    shots_rows.sort(key=lambda r: r["CutOrder"])
    elements_rows.sort(key=lambda r: (r["CutOrder"]))

    return shots_rows, elements_rows

def compare_edits(current_shots, old_shots):

    for current_shot in current_shots:
        for old_shot in old_shots:
            if old_shot['ShotCode'] == current_shot['ShotCode']:
                d_in = current_shot['CutIn'] - old_shot['CutIn']
                d_out = current_shot['CutOut'] - old_shot['CutOut']
                change_to_cut = ""
                if d_in != 0 or d_out != 0:
                    change_to_cut = f"In: {d_in}, Out: {d_out}" if (d_in != 0 and d_out != 0) \
                    else f"In:{d_in}" if d_in != 0 \
                    else f"Out:{d_out}"
                if current_shot['RetimeSummary'] != old_shot['RetimeSummary']:
                    change_to_cut += "retime change" if (change_to_cut == "") \
                    else ", retime change"
                current_shot['ChangeToCut'] = change_to_cut

# -------- Main extraction --------

def main():
    args = parse_args()
    app = resolve_app()
    project = get_project(app)
    timeline = get_timeline(project, args.timeline)
    if args.old_timeline:
        old_timeline = get_timeline(project, args.old_timeline)
    fps = get_timeline_fps(timeline)

    current_shots, current_elements = read_edit(args, app, project, timeline, fps)
    if args.old_timeline:
        old_shots, old_elements = read_edit(args, app, project, old_timeline, fps)
        compare_edits(current_shots, old_shots)

    # -------- Excel output --------
    wb = Workbook()
    # Shots sheet
    ws_shots = wb.active
    ws_shots.title = "Shots"
    shots_cols = [
        "Sequence", "Cut Order", "Editorial Name", "Shot Code", "Change to Cut",
        "Work In", "Cut In", "Cut Out", "Work Out",
        "Cut Duration", "Bg Retime", "Fg Retime", "Cut In TC"
    ]
    ws_shots.append(shots_cols)
    for r in current_shots:
        ws_shots.append([
            r["Sequence"], r["CutOrder"], r["EditorialName"], r["ShotCode"], r["ChangeToCut"],
            r["WorkIn"], r["CutIn"], r["CutOut"], r["WorkOut"],
            r["CutDuration"], r["BgRetime"], r["FgRetime"], r["CutInTC"]
        ])

    # Elements sheet
    ws_elems = wb.create_sheet(title="Elements")
    elems_cols = [
        "Sequence", "Cut Order", "Editorial Name", "Shot Code", "Element",
        "Cut In", "Cut Out",
        "Clip In TC", "Clip In Frames", "Clip In", "Clip Out", "Clip Out Frames", "Clip Out TC", "Clip Duration", 
        "Retime Summary", "Scale & Repo"
    ]
    ws_elems.append(elems_cols)
    for r in current_elements:
        ws_elems.append([
            r["Sequence"], r["CutOrder"], r["EditorialName"], r["ShotCode"], r["Element"],
            r["ShotCutIn"], r["ShotCutOut"],
            r["ClipInTC"], r["ClipInFrames"], r["ClipIn"], r["ClipOut"], r["ClipOutFrames"], r["ClipOutTC"], r['ClipDuration'],
            r["Retime"], r["ScaleRepo"]
        ])

    # Auto-width (simple)
    for ws in (ws_shots, ws_elems):
        for col_idx, _ in enumerate(ws[1], start=1):
            max_len = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                val = row[0].value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 50)
    if args.output != None:
        out_path = os.path.abspath(args.output)
    else:
        out_path = os.path.abspath(timeline_name + '_shot_list.xlsx')
    wb.save(out_path)
    print(f"✓ Wrote Excel: {out_path}")

if __name__ == "__main__":
    main()
