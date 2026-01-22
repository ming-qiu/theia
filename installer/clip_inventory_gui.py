"""
Theia - Clip Inventory GUI
Exports DaVinci Resolve timeline clips to Excel with thumbnails
"""
import sys
import base64
import time
from io import BytesIO
from pathlib import Path
from timecode import Timecode

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QComboBox, QFileDialog,
    QMessageBox, QProgressBar, QTextEdit
)
from PySide6.QtCore import Qt, QThread, Signal, QUrl
from PySide6.QtGui import QFont, QDesktopServices, QIcon

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

# Import DaVinci Resolve API
try:
    import DaVinciResolveScript as dvr
except ImportError:
    # If not in path, try to add it
    import sys
    resolve_script_api = "/Library/Application Support/Blackmagic Design/DaVinci Resolve/Developer/Scripting/Modules"
    if resolve_script_api not in sys.path:
        sys.path.append(resolve_script_api)
    try:
        import DaVinciResolveScript as dvr
    except ImportError:
        dvr = None


class ExportWorker(QThread):
    """Threaded export worker to keep GUI responsive."""
    progress = Signal(str)
    finished = Signal(bool, str, int)  # success, message, clip_count
    
    def __init__(self, output_path, bg_track):
        super().__init__()
        self.output_path = output_path
        self.bg_track = bg_track
    
    def log(self, msg):
        self.progress.emit(msg)
    
    def get_thumbnail(self, timeline, clip, fps):
        """Get clip thumbnail from timeline."""
        try:
            clip_start = clip.GetStart() + 1
            tc = Timecode(fps, frames=int(clip_start))
            
            if not timeline.SetCurrentTimecode(str(tc)):
                return None
            
            thumb = timeline.GetCurrentClipThumbnailImage()
            if not thumb or not thumb.get('data'):
                return None
            
            img_bytes = base64.b64decode(thumb['data'])
            img = PILImage.frombytes('RGB', (thumb['width'], thumb['height']), img_bytes)
            
            # Resize to standard height
            aspect = img.width / img.height
            new_h = 150
            new_w = int(new_h * aspect)
            return img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)
            
        except Exception as e:
            self.log(f"    Thumbnail error: {e}")
            return None
    
    def run(self):
        """Main export logic."""
        try:
            # Connect to Resolve
            self.log("Connecting to DaVinci Resolve...")
            
            if dvr is None:
                self.finished.emit(False, "DaVinci Resolve API not available. Make sure Resolve is running.", 0)
                return
            
            resolve = dvr.scriptapp("Resolve")
            if not resolve:
                self.finished.emit(False, "Could not connect to Resolve. Make sure DaVinci Resolve is running.", 0)
                return
            
            project = resolve.GetProjectManager().GetCurrentProject()
            timeline = project.GetCurrentTimeline() if project else None
            
            if not timeline:
                self.finished.emit(False, "No timeline open", 0)
                return
            
            # Save current state
            starting_page = resolve.GetCurrentPage()
            starting_timecode = timeline.GetCurrentTimecode()
            self.log(f"Saving playhead position: {starting_timecode}")
            
            # Switch to Color page for thumbnails
            self.log("Opening Color page...")
            resolve.OpenPage('color')
            time.sleep(0.5)
            
            # Get timeline info
            timeline_name = timeline.GetName()
            fps = float(timeline.GetSetting("timelineFrameRate"))
            self.log(f"Timeline: {timeline_name} ({fps} fps)")
            
            # Get clips
            clips = timeline.GetItemListInTrack("video", self.bg_track)
            if not clips:
                resolve.OpenPage(starting_page)
                timeline.SetCurrentTimecode(starting_timecode)
                self.finished.emit(False, f"No clips on track {self.bg_track}", 0)
                return
            
            self.log(f"Found {len(clips)} clips\n")
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Shots"
            
            # Headers
            headers = ["Thumbnail", "Reel Name", "Cut Order", "Record In", 
                      "Record Out", "Source In", "Notes"]
            for idx, header in enumerate(headers, 1):
                ws.cell(1, idx, header)
            
            # Column widths
            ws.column_dimensions['A'].width = 34
            ws.column_dimensions['B'].width = 25
            for col in ['C', 'D', 'E', 'F', 'G']:
                ws.column_dimensions[col].width = 15
            
            # Process clips
            for idx, clip in enumerate(clips, 1):
                self.log(f"[{idx}/{len(clips)}] {clip.GetName()}")
                row = idx + 1
                
                # Basic info
                ws.cell(row, 2, clip.GetName())
                ws.cell(row, 3, idx)
                
                # Timecodes
                tc_in = Timecode(fps, frames=int(clip.GetStart() + 1))
                tc_out = Timecode(fps, frames=int(clip.GetEnd() + 1))
                ws.cell(row, 4, str(tc_in))
                ws.cell(row, 5, str(tc_out))
                
                # Source timecode
                media_item = clip.GetMediaPoolItem()
                if media_item:
                    try:
                        props = media_item.GetClipProperty()
                        src_fps = float(props.get('FPS', fps))
                        src_tc_str = props.get('Start TC')
                        src_start = Timecode(str(src_fps), src_tc_str).frames
                        src_tc = Timecode(src_fps, frames=int(src_start + clip.GetLeftOffset()))
                        ws.cell(row, 6, str(src_tc))
                    except:
                        ws.cell(row, 6, str(tc_in))
                else:
                    ws.cell(row, 6, str(tc_in))
                
                # Thumbnail
                if media_item:
                    thumb = self.get_thumbnail(timeline, clip, fps)
                    if thumb:
                        buf = BytesIO()
                        thumb.save(buf, format='PNG')
                        buf.seek(0)
                        ws.add_image(XLImage(buf), f'A{row}')
                        ws.row_dimensions[row].height = 112.5
                    else:
                        ws.cell(row, 1, "No thumbnail")
                else:
                    ws.cell(row, 1, "Generator")
            
            # Save
            self.log(f"\nSaving to {self.output_path}...")
            wb.save(self.output_path)
            
            # Restore timeline state
            self.log(f"Restoring playhead to: {starting_timecode}")
            timeline.SetCurrentTimecode(starting_timecode)
            resolve.OpenPage(starting_page)
            
            self.finished.emit(True, f"Successfully exported {len(clips)} clips", len(clips))
            
        except Exception as e:
            self.log(f"\nERROR: {e}")
            self.finished.emit(False, str(e), 0)


class ClipInventoryGUI(QMainWindow):
    """Main GUI window for Clip Inventory."""
    
    def __init__(self):
        super().__init__()
        self.worker = None
        self.last_export_path = None
        self.load_resources()
        self.setup_ui()
        self.populate_track_list()
    
    def load_resources(self):
        """Load icon and stylesheet from Theia resources folder."""
        theia_dir = Path("/Library/Application Support/Theia")
        
        # Load icon
        icon_path = theia_dir / "resources" / "icon.png"
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))
        
        # Load stylesheet
        style_path = theia_dir / "resources" / "win95.qss"
        if style_path.exists():
            with open(style_path, 'r') as f:
                self.setStyleSheet(f.read())
    
    def setup_ui(self):
        """Build the interface."""
        self.setWindowTitle("Theia - Clip Inventory")
        self.setMinimumSize(600, 500)
        
        # Main layout
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Title
        title = QLabel("Export Timeline to Excel")
        font = QFont()
        font.setPointSize(16)
        font.setBold(True)
        title.setFont(font)
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Output file
        file_row = QHBoxLayout()
        file_row.addWidget(QLabel("Output File:"))
        self.file_input = QLineEdit(str(Path.home() / "Downloads/clip_inventory.xlsx"))
        file_row.addWidget(self.file_input)
        browse = QPushButton("Browse...")
        browse.clicked.connect(self.browse_file)
        file_row.addWidget(browse)
        layout.addLayout(file_row)
        
        # Track number
        track_row = QHBoxLayout()
        track_row.addWidget(QLabel("Video Track:"))
        self.track_combo = QComboBox()
        self.track_combo.setMinimumWidth(100)
        track_row.addWidget(self.track_combo)
        
        # Refresh button to reload track list
        refresh_btn = QPushButton("↻")
        refresh_btn.setMaximumWidth(40)
        refresh_btn.setToolTip("Refresh track list")
        refresh_btn.clicked.connect(self.populate_track_list)
        track_row.addWidget(refresh_btn)
        
        track_row.addStretch()
        layout.addLayout(track_row)
        
        # Export button
        self.export_btn = QPushButton("Export")
        self.export_btn.setMinimumHeight(40)
        self.export_btn.clicked.connect(self.start_export)
        layout.addWidget(self.export_btn)
        
        # Progress bar
        self.progress = QProgressBar()
        self.progress.hide()
        layout.addWidget(self.progress)
        
        # Log
        layout.addWidget(QLabel("Log:"))
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        layout.addWidget(self.log)
    
    def populate_track_list(self):
        """Get available video tracks from current timeline."""
        self.track_combo.clear()
        
        try:
            if dvr is None:
                self.track_combo.addItem("Track 1", 1)
                return
            
            resolve = dvr.scriptapp("Resolve")
            if not resolve:
                self.track_combo.addItem("Track 1", 1)
                return
            
            project = resolve.GetProjectManager().GetCurrentProject()
            timeline = project.GetCurrentTimeline() if project else None
            
            if not timeline:
                self.track_combo.addItem("Track 1", 1)
                self.log.append("⚠️  No timeline open - defaulting to track 1")
                return
            
            # Get track count
            track_count = timeline.GetTrackCount("video")
            
            if track_count == 0:
                self.track_combo.addItem("Track 1", 1)
                self.log.append("⚠️  No video tracks found - defaulting to track 1")
                return
            
            # Populate combo with available tracks
            for i in range(1, track_count + 1):
                self.track_combo.addItem(f"Track {i}", i)
            
            self.log.append(f"✓ Found {track_count} video track(s)")
            
        except Exception as e:
            self.track_combo.addItem("Track 1", 1)
            self.log.append(f"⚠️  Could not get tracks: {e}")
    
    def browse_file(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Inventory", 
            self.file_input.text(),
            "Excel Files (*.xlsx)"
        )
        if path:
            self.file_input.setText(path)
    
    def start_export(self):
        """Begin export process."""
        output = self.file_input.text()
        if not output:
            QMessageBox.warning(self, "Error", "Please specify output file")
            return
        
        # Get selected track number
        track = self.track_combo.currentData()
        if track is None:
            track = 1
        
        self.log.clear()
        self.export_btn.setEnabled(False)
        self.progress.setRange(0, 0)
        self.progress.show()
        
        # Store path for later
        self.last_export_path = output
        
        self.worker = ExportWorker(output, track)
        self.worker.progress.connect(self.update_log)
        self.worker.finished.connect(self.export_done)
        self.worker.start()
    
    def update_log(self, msg):
        self.log.append(msg)
        self.log.verticalScrollBar().setValue(
            self.log.verticalScrollBar().maximum()
        )
    
    def export_done(self, success, msg, clip_count):
        """Handle export completion."""
        self.export_btn.setEnabled(True)
        self.progress.hide()
        
        if success:
            # Create custom message box with "Open File" button
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Export Complete")
            msg_box.setText(f"Successfully exported {clip_count} clip(s)")
            msg_box.setIcon(QMessageBox.Information)
            
            # Add custom buttons
            open_btn = msg_box.addButton("Open File", QMessageBox.ActionRole)
            close_btn = msg_box.addButton(QMessageBox.Ok)
            
            msg_box.exec()
            
            # Check which button was clicked
            if msg_box.clickedButton() == open_btn:
                self.open_export_file()
        else:
            QMessageBox.critical(self, "Error", f"Export failed: {msg}")
    
    def open_export_file(self):
        """Open the exported Excel file in default application."""
        if self.last_export_path and Path(self.last_export_path).exists():
            QDesktopServices.openUrl(QUrl.fromLocalFile(self.last_export_path))
        else:
            QMessageBox.warning(self, "Error", "Export file not found")


def main():
    app = QApplication(sys.argv)
    window = ClipInventoryGUI()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()