"""
DaVinci Resolve - Export Timeline Cut Points to Excel
Exports cut information including thumbnails to an Excel spreadsheet.
"""

import sys
import base64
import time
from io import BytesIO
from timecode import Timecode
import time
import argparse

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl package required. Install with: pip install openpyxl")
    sys.exit(1)

try:
    from PIL import Image as PILImage
except ImportError:
    print("Error: Pillow package required. Install with: pip install Pillow")
    sys.exit(1)

# Import DaVinci Resolve API
try:
    import DaVinciResolveScript as dvr
except ImportError:
    print("Error: Could not import DaVinci Resolve API")
    sys.exit(1)


def parse_args():
    parser = argparse.ArgumentParser(description="Create an inventory of BG clips")
    parser.add_argument("--file-name", default = "clip_inventory.xlsx", help="Output file name")
    parser.add_argument("--bg-track", default = 1, help = "Track number of the BG track")
    return parser.parse_args()

def get_resolve_objects():
    """Initialize and return Resolve API objects."""
    resolve = dvr.scriptapp("Resolve")
    starting_page = resolve.GetCurrentPage()

    if not resolve:
        print("Error: Could not connect to DaVinci Resolve")
        sys.exit(1)
    
    pm = resolve.GetProjectManager()
    project = pm.GetCurrentProject()
    if not project:
        print("Error: No project is currently open")
        sys.exit(1)
    
    timeline = project.GetCurrentTimeline()
    if not timeline:
        print("Error: No timeline is currently open")
        sys.exit(1)

    try:
        print ("Opening Color page...")
        resolve.OpenPage('color')
        time.sleep(1)
    except:
        print ("Error: Cannot go to Color page")
        sys.exit(1)
    
    return resolve, project, timeline, resolve, starting_page


def get_clip_thumbnail(resolve, project, timeline, clip, timeline_fps):
    """
    Get thumbnail for a clip by setting playhead and grabbing current frame.
    Returns PIL Image object or None.
    """
    try:
        # Get clip's record start position
        clip_start = clip.GetStart() + 1
        
        # Convert to timecode and set playhead
        start_tc = Timecode(timeline_fps, frames=int(clip_start))
        
        print(f"    Setting playhead to: {start_tc}")
        
        # Set the timeline playhead to the clip's start
        if not timeline.SetCurrentTimecode(str(start_tc)):
            print(f"    Warning: Could not set timecode to {start_tc}")
            return None
        
        # Grab the thumbnail
        print(f"    Attempting to grab thumbnail...")
        thumb_data = timeline.GetCurrentClipThumbnailImage()
        thumb_data = timeline.GrabStill()
        
        if not thumb_data:
            print(f"    Warning: GetCurrentClipThumbnailImage returned None")
            return None
        
        print(f"    Thumbnail data received: {type(thumb_data)}")
        print(f"    Keys: {thumb_data.keys() if isinstance(thumb_data, dict) else 'Not a dict'}")
        
        # Decode the base64 image data
        width = thumb_data.get('width', 0)
        height = thumb_data.get('height', 0)
        img_format = thumb_data.get('format', '')
        img_data_b64 = thumb_data.get('data', '')
        
        print(f"    Width: {width}, Height: {height}, Format: {img_format}")
        print(f"    Data length: {len(img_data_b64) if img_data_b64 else 0}")
        
        if not img_data_b64:
            print(f"    Warning: No image data in thumbnail")
            return None
        
        # Decode base64 to bytes
        img_bytes = base64.b64decode(img_data_b64)
        print(f"    Decoded {len(img_bytes)} bytes")
        
        # Create PIL Image from raw RGB data
        img = PILImage.frombytes('RGB', (width, height), img_bytes)
        print(f"    Successfully created PIL Image")
        return img
        
    except Exception as e:
        print(f"    Exception: Error getting thumbnail: {e}")
        import traceback
        traceback.print_exc()
        return None


def export_cuts_to_excel(output_path, bg_track):
    """Main function to export timeline cuts to Excel."""
    
    print("Connecting to DaVinci Resolve...")
    resolve, project, timeline, resolve, starting_page = get_resolve_objects()
    
    timeline_name = timeline.GetName()
    timeline_fps = float(timeline.GetSetting("timelineFrameRate"))
    timeline_start_frame = timeline.GetStartFrame()
    
    print(f"Timeline: {timeline_name}")
    print(f"FPS: {timeline_fps}")
    print(f"Start Frame: {timeline_start_frame}")
    
    # Get all clips from video track 1
    print(f"\nGetting clips from video track {bg_track}...")
    clips = timeline.GetItemListInTrack("video", bg_track)
    
    if not clips:
        print(f"Error: No clips found on video track {bg_track}")
        return
    
    print(f"Found {len(clips)} clips")
    
    # Create workbook
    print("\nCreating Excel workbook...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Cut Points"
    
    # Set column headers
    ws['A1'] = "Thumbnail"
    ws['B1'] = "Reel Name"
    ws['C1'] = "Cut Order"
    ws['D1'] = "Record TC In"
    ws['E1'] = "Record TC Out"
    ws['F1'] = "Source TC In"
    ws['G1'] = "VFX Shot Code"
    ws['H1'] = "VFX Work"
    ws['I1'] = "Vendor"
    
    # Set column widths
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    
    # Process each clip
    for idx, clip in enumerate(clips, start=1):
        print(f"\nProcessing clip {idx}/{len(clips)}...")
        row = idx + 1  # +1 for header row
        
        # C: Cut Order
        ws[f'C{row}'] = idx
        
        # D: Record TC In (timeline position of clip start)
        clip_start_frame = clip.GetStart() + 1
        record_tc_in = Timecode(timeline_fps, frames=int(clip_start_frame))
        ws[f'D{row}'] = str(record_tc_in)
        print(f"  Record TC In: {record_tc_in}")
        
        # E: Record TC Out (timeline position of clip end)
        clip_end_frame = clip.GetEnd() + 1
        record_tc_out = Timecode(timeline_fps, frames=int(clip_end_frame))
        ws[f'E{row}'] = str(record_tc_out)
        print(f"  Record TC Out: {record_tc_out}")
        
        # F: Source TC (in point of the clip in source media)
        # Get source FPS from media pool item if available
        media_pool_item = clip.GetMediaPoolItem()
        try:
            clip_props = media_pool_item.GetClipProperty()
            source_fps_str = clip_props.get('FPS', str(timeline_fps))
            source_fps = float(source_fps_str)
        except:
            source_fps = timeline_fps
        
        take_start_tc = media_pool_item.GetClipProperty().get('Start TC')
        take_start_frame = Timecode(str(source_fps), take_start_tc).frames
        left_offset = clip.GetLeftOffset()       
        clip_start_frame = take_start_frame + left_offset

        source_tc = Timecode(source_fps, frames=int(clip_start_frame))
        ws[f'F{row}'] = str(source_tc)
        print(f"  Source TC: {source_tc}")
        
        # B: Reel Name
        reel_name = clip.GetName()
        ws[f'B{row}'] = reel_name
        print(f"  Reel Name: {reel_name}")
        
        # A: Thumbnail
        print(f"  Getting thumbnail...")
        thumb_img = get_clip_thumbnail(resolve, project, timeline, clip, timeline_fps)
        
        if thumb_img:
            # Resize thumbnail to reasonable size (max 150 pixels height)
            max_height = 150
            aspect_ratio = thumb_img.width / thumb_img.height
            new_height = max_height
            new_width = int(new_height * aspect_ratio)
            thumb_img = thumb_img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
            
            # Save to BytesIO and insert into Excel
            img_buffer = BytesIO()
            thumb_img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            xl_img = XLImage(img_buffer)
            ws.add_image(xl_img, f'A{row}')
            
            # Adjust row height to fit thumbnail
            ws.row_dimensions[row].height = max_height * 0.75  # Excel points conversion
            print(f"  Thumbnail added")
        else:
            ws[f'A{row}'] = "No thumbnail"
            print(f"  No thumbnail available")
    
    # Save workbook
    print(f"\nSaving Excel file to: {output_path}")
    wb.save(output_path)
    print("Export complete!")

    resolve.OpenPage(starting_page)


if __name__ == "__main__":
    # Default output path - modify as needed
    args = parse_args()
    output_file = args.file_name
    bg_track = int(args.bg_track)
    
    export_cuts_to_excel(output_file, bg_track)