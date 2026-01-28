"""
Theia - Shot Metadata GUI
Import VFX shot codes from Excel to timeline subtitles and add frame counters
"""

import sys
import os
from pathlib import Path
from timecode import Timecode

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QComboBox, QFileDialog,
    QMessageBox, QProgressBar, QTextEdit, QCheckBox, QGroupBox,
    QSpinBox, QDoubleSpinBox, QScrollArea
)
from PySide6.QtCore import Qt, QThread, Signal, QUrl
from PySide6.QtGui import QFont, QDesktopServices, QIcon

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

# Import DaVinci Resolve API
try:
    import DaVinciResolveScript as dvr
except ImportError:
    import sys
    resolve_script_api = "/Library/Application Support/Blackmagic Design/DaVinci Resolve/Developer/Scripting/Modules"
    if resolve_script_api not in sys.path:
        sys.path.append(resolve_script_api)
    try:
        import DaVinciResolveScript as dvr
    except ImportError:
        dvr = None


class MetadataWorker(QThread):
    """Threaded worker for processing metadata and frame counters."""
    progress = Signal(str)
    finished = Signal(bool, str)
    
    def __init__(self, sheet_path, selected_columns, output_dir, fps,
                 frame_counter_path=None, first_frame=None):
        super().__init__()
        self.sheet_path = sheet_path
        self.selected_columns = selected_columns  # List of (column_index, column_name) tuples
        self.output_dir = output_dir
        self.fps = fps
        self.frame_counter_path = frame_counter_path
        self.first_frame = first_frame
    
    def log(self, msg):
        self.progress.emit(msg)
    
    def create_srt_file(self, ws, output_path, column_index):
        """Create SRT file from specific Excel column."""
        
        # Read data from Excel
        subtitles = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= column_index:
                continue
            
            content = row[column_index]
            if content and str(content).strip():
                subtitles.append({
                    'tc_in': Timecode(self.fps, str(row[3])),    # Column D
                    'tc_out': Timecode(self.fps, str(row[4])),   # Column E
                    'text': str(content).strip()
                })
        
        if not subtitles:
            return None
        
        # Create SRT content
        srt_lines = []
        for idx, sub in enumerate(subtitles, start=1):
            # Convert to SRT format: HH:MM:SS,mmm
            def to_srt(tc):
                total_sec = tc.frames / float(tc.framerate)
                h = int(total_sec // 3600)
                m = int((total_sec % 3600) // 60)
                s = int(total_sec % 60)
                ms = int((total_sec % 1) * 1000)
                return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"
            
            srt_lines.append(str(idx))
            srt_lines.append(f"{to_srt(sub['tc_in'])} --> {to_srt(sub['tc_out'])}")
            srt_lines.append(sub['text'])
            srt_lines.append("")
        
        # Write SRT file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(srt_lines))
        
        return output_path
    
    def add_frame_counters(self):
        """Add frame counter videos to timeline based on shot timings from Excel."""
        
        if not dvr:
            self.log("ERROR: DaVinci Resolve API not available")
            return False
        
        try:
            # Connect to Resolve
            resolve = dvr.scriptapp("Resolve")
            project = resolve.GetProjectManager().GetCurrentProject()
            timeline = project.GetCurrentTimeline()
            mediapool = project.GetMediaPool()
            
            self.log(f"Timeline: {timeline.GetName()}")
            
            # Import frame counter to media pool
            abs_frame_counter_path = os.path.abspath(self.frame_counter_path)
            self.log(f"Importing frame counter: {abs_frame_counter_path}")
            imported = mediapool.ImportMedia([abs_frame_counter_path])
            if not imported:
                self.log("ERROR: Failed to import frame counter video")
                return False
            
            frame_counter_item = imported[0]
            fc_first_frame = Timecode(self.fps, frame_counter_item.GetClipProperty("Start TC")).frames - 1
            
            # Get number of video tracks and create new track
            num_tracks = timeline.GetTrackCount("video")
            target_track = num_tracks + 1
            self.log(f"Adding frame counters to track {target_track}")
            timeline.AddTrack("video", None)
            
            # Load Excel to get shot timings
            wb = load_workbook(self.sheet_path)
            ws = wb.active
            
            # Read shot data
            clips_to_add = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) < 5:
                    continue
                
                # Check if any metadata exists in column G onwards (index 6+)
                has_metadata = False
                for cell_value in row[6:]:  # Column G is index 6
                    if cell_value and str(cell_value).strip():
                        has_metadata = True
                        break
                
                # Skip this shot if no metadata
                if not has_metadata:
                    continue
                
                record_tc_in = Timecode(self.fps, str(row[3]))   # Column D
                record_tc_out = Timecode(self.fps, str(row[4]))  # Column E
                
                shot_duration = record_tc_out.frames - record_tc_in.frames
                
                clips_to_add.append({
                    "mediaPoolItem": frame_counter_item,
                    "startFrame": self.first_frame - fc_first_frame,
                    "endFrame": self.first_frame - fc_first_frame + shot_duration,
                    "trackIndex": target_track,
                    "recordFrame": record_tc_in.frames - 1
                })
            
            self.log(f"Adding {len(clips_to_add)} frame counter clips...")
            result = mediapool.AppendToTimeline(clips_to_add)
            
            if result:
                self.log(f"Success! Added {len(result)} frame counter clips to track {target_track}")
                return True
            else:
                self.log("ERROR: Failed to add clips to timeline")
                return False
                
        except Exception as e:
            self.log(f"ERROR: {str(e)}")
            return False
    
    def run(self):
        """Execute the metadata export and/or frame counter addition."""
        try:
            success_count = 0
            
            # Handle frame counter mode
            if self.frame_counter_path and self.first_frame is not None:
                self.log("=" * 50)
                self.log("Adding Frame Counters")
                self.log("=" * 50)
                if self.add_frame_counters():
                    success_count += 1
                else:
                    self.finished.emit(False, "Frame counter addition failed")
                    return
            
            # Create SRT files
            if self.selected_columns:
                self.log("=" * 50)
                self.log("Creating SRT Subtitle Files")
                self.log("=" * 50)
                
                # Load sheet
                wb = load_workbook(self.sheet_path)
                ws = wb.active
                
                # Create SRT files for each selected column
                srt_count = 0
                for col_idx, col_name in self.selected_columns:
                    # Create output filename from header
                    output_path = os.path.join(self.output_dir, f"{col_name}.srt")
                    if self.create_srt_file(ws, output_path, col_idx):
                        self.log(f"  Created: {output_path}")
                        srt_count += 1
                
                if srt_count > 0:
                    self.log(f"Successfully created {srt_count} SRT file(s)")
                    success_count += 1
                else:
                    self.log("No SRT files created (no data found in specified columns)")
            
            if success_count > 0:
                self.finished.emit(True, "Processing completed successfully")
            else:
                self.finished.emit(False, "No operations completed")
                
        except Exception as e:
            self.log(f"ERROR: {str(e)}")
            self.finished.emit(False, str(e))


class AddMetadataGUI(QMainWindow):
    """Main GUI window for shot metadata operations."""
    
    def __init__(self):
        super().__init__()
        self.last_export_path = None
        self.setup_ui()

    def setup_ui(self):
        self.setWindowTitle("Theia - Add Metadata")
        self.setMinimumWidth(750)
        self.setMinimumHeight(700)
        
        self.column_checkboxes = []
        self.available_columns = []  # List of (index, name) tuples
        
        # Create central widget
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Title
        title = QLabel("Add Metadata to Timeline")
        font = QFont()
        font.setPointSize(16)
        font.setBold(True)
        title.setFont(font)
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        layout.addSpacing(5)
        
        # Excel file input
        file_group = QGroupBox("Excel File")
        file_layout = QVBoxLayout()
        
        file_row = QHBoxLayout()
        file_row.addWidget(QLabel("Metadata File:"))
        self.sheet_input = QLineEdit(str(Path.home() / "Downloads/clip_inventory.xlsx"))
        self.sheet_input.textChanged.connect(self.on_excel_file_changed)
        file_row.addWidget(self.sheet_input)
        browse_sheet_btn = QPushButton("Browse...")
        browse_sheet_btn.clicked.connect(self.browse_sheet)
        file_row.addWidget(browse_sheet_btn)
        file_layout.addLayout(file_row)
        file_group.setLayout(file_layout)
        layout.addWidget(file_group)
        layout.addSpacing(5)
        
        # SRT Subtitle Settings
        srt_group = QGroupBox("SRT Subtitle Files")
        srt_layout = QVBoxLayout()
        
        # Column selection header
        col_header = QHBoxLayout()
        col_header.addWidget(QLabel("Select Metadata to Export:"))
        
        select_all_btn = QPushButton("Select All")
        select_all_btn.setMaximumWidth(80)
        select_all_btn.clicked.connect(self.select_all_columns)
        col_header.addWidget(select_all_btn)
        
        deselect_all_btn = QPushButton("Deselect All")
        deselect_all_btn.setMaximumWidth(100)
        deselect_all_btn.clicked.connect(self.deselect_all_columns)
        col_header.addWidget(deselect_all_btn)
        
        refresh_btn = QPushButton("↻")
        refresh_btn.setMaximumWidth(40)
        refresh_btn.setToolTip("Refresh Column List")
        refresh_btn.clicked.connect(self.load_excel_columns)
        col_header.addWidget(refresh_btn)
        
        col_header.addStretch()
        srt_layout.addLayout(col_header)
        
        # Scrollable checkbox area for columns
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMaximumHeight(150)
        scroll.setMinimumHeight(60)
        
        self.column_checkbox_widget = QWidget()
        self.column_checkbox_layout = QVBoxLayout(self.column_checkbox_widget)
        self.column_checkbox_layout.setContentsMargins(5, 5, 5, 5)
        
        scroll.setWidget(self.column_checkbox_widget)
        srt_layout.addWidget(scroll)
        
        # Output directory selection
        output_row = QHBoxLayout()
        output_row.addWidget(QLabel("SRT Output Directory:"))
        self.output_dir_input = QLineEdit(str(Path.home() / "Downloads/"))
        self.output_dir_input.setPlaceholderText("Directory for SRT files")
        output_row.addWidget(self.output_dir_input)
        
        browse_output_btn = QPushButton("Browse...")
        browse_output_btn.setMaximumWidth(100)
        browse_output_btn.clicked.connect(self.browse_output_dir)
        output_row.addWidget(browse_output_btn)
        
        srt_layout.addLayout(output_row)
        
        info_label = QLabel("ℹ️  SRT files will be named after column headers")
        info_label.setStyleSheet("color: #666; font-size: 10px;")
        srt_layout.addWidget(info_label)
        
        srt_group.setLayout(srt_layout)
        layout.addWidget(srt_group)
        layout.addSpacing(5)
        
        # Frame Counter Settings
        fc_group = QGroupBox("Frame Counter (Optional)")
        fc_layout = QVBoxLayout()
        
        # Enable checkbox
        self.fc_enable = QCheckBox("Add frame counter videos to timeline")
        self.fc_enable.toggled.connect(self.toggle_frame_counter)
        fc_layout.addWidget(self.fc_enable)
        
        # Frame counter file
        fc_file_row = QHBoxLayout()
        fc_file_row.addWidget(QLabel("Frame Counter File:"))
        self.fc_file_input = QLineEdit()
        self.fc_file_input.setPlaceholderText("Path to frame counter video file")
        self.fc_file_input.setEnabled(False)
        fc_file_row.addWidget(self.fc_file_input)
        
        browse_fc_btn = QPushButton("Browse...")
        browse_fc_btn.setMaximumWidth(100)
        browse_fc_btn.clicked.connect(self.browse_frame_counter)
        browse_fc_btn.setEnabled(False)
        self.browse_fc_btn = browse_fc_btn
        fc_file_row.addWidget(browse_fc_btn)
        
        fc_layout.addLayout(fc_file_row)
        
        # First frame
        first_frame_row = QHBoxLayout()
        first_frame_row.addWidget(QLabel("Starting Frame Number:"))
        self.first_frame = QSpinBox()
        self.first_frame.setMinimum(0)
        self.first_frame.setMaximum(999999)
        self.first_frame.setValue(1001)
        self.first_frame.setEnabled(False)
        first_frame_row.addWidget(self.first_frame)
        first_frame_row.addStretch()
        
        fc_layout.addLayout(first_frame_row)
        
        fc_group.setLayout(fc_layout)
        layout.addWidget(fc_group)
        layout.addSpacing(5)
        
        # Frame Rate Settings
        fps_group = QGroupBox("Set Frame Rate")
        fps_layout = QVBoxLayout()
        
        fps_row = QHBoxLayout()
        fps_row.addWidget(QLabel("Timeline FPS:"))
        
        self.fps_combo = QComboBox()
        self.fps_combo.setMaximumWidth(120)
        self.fps_combo.addItems([
            "23.976",
            "24",
            "25",
            "30",
            "60",
            "Custom..."
        ])
        self.fps_combo.setCurrentText("24")
        self.fps_combo.currentTextChanged.connect(self.on_fps_changed)
        fps_row.addWidget(self.fps_combo)
        
        # Custom FPS input (hidden by default)
        self.custom_fps_input = QLineEdit()
        self.custom_fps_input.setMaximumWidth(100)
        self.custom_fps_input.setPlaceholderText("Enter FPS")
        self.custom_fps_input.hide()
        self.custom_fps_input.textChanged.connect(self.validate_custom_fps)
        fps_row.addWidget(self.custom_fps_input)
        
        fps_row.addStretch()
        
        fps_layout.addLayout(fps_row)
        fps_group.setLayout(fps_layout)
        layout.addWidget(fps_group)
        layout.addSpacing(5)
        
        # Add Metadata button
        self.process_btn = QPushButton("Export Metadata to SRT")
        self.process_btn.setMinimumHeight(40)
        self.process_btn.clicked.connect(self.start_processing)
        layout.addWidget(self.process_btn)
        
        # Progress bar
        self.progress = QProgressBar()
        self.progress.hide()
        layout.addWidget(self.progress)
        
        # Log
        layout.addWidget(QLabel("Log:"))
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        layout.addWidget(self.log)
        
        # Check if Resolve is available
        self.check_resolve_connection()
        
        # Try to load columns from default file
        self.load_excel_columns()
    
    def on_fps_changed(self, text):
        """Handle FPS combo box changes."""
        if text == "Custom...":
            self.custom_fps_input.show()
            self.custom_fps_input.setFocus()
        else:
            self.custom_fps_input.hide()
    
    def validate_custom_fps(self, text):
        """Validate custom FPS input."""
        if not text:
            return
        
        try:
            fps = float(text)
            if fps <= 0:
                self.custom_fps_input.setStyleSheet("background-color: #ffcccc;")
            else:
                self.custom_fps_input.setStyleSheet("")
        except ValueError:
            self.custom_fps_input.setStyleSheet("background-color: #ffcccc;")
    
    def get_fps(self):
        """Get the selected FPS value."""
        if self.fps_combo.currentText() == "Custom...":
            try:
                fps = float(self.custom_fps_input.text())
                if fps <= 0:
                    return None
                return fps
            except ValueError:
                return None
        else:
            return float(self.fps_combo.currentText())
    
    def on_excel_file_changed(self):
        """Handle Excel file path changes."""
        # Auto-load columns when a valid file is entered
        if os.path.exists(self.sheet_input.text()):
            self.load_excel_columns()
    
    def load_excel_columns(self):
        """Load column headers from Excel file starting from column G."""
        # Clear existing checkboxes
        for cb in self.column_checkboxes:
            cb.deleteLater()
        self.column_checkboxes.clear()
        self.available_columns.clear()
        
        sheet_path = self.sheet_input.text()
        if not sheet_path or not os.path.exists(sheet_path):
            self.add_column_checkbox(6, "G", "(No file loaded)", enabled=False)
            return
        
        try:
            wb = load_workbook(sheet_path, read_only=True)
            ws = wb.active
            
            # Read headers from first row, starting at column G (index 7, 0-indexed)
            headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            
            found_columns = False
            for col_idx in range(6, len(headers)):  # Start from column G (index 6)
                header = headers[col_idx]
                if header and str(header).strip():
                    col_letter = get_column_letter(col_idx + 1)
                    col_name = str(header).strip()
                    self.available_columns.append((col_idx, col_name))
                    self.add_column_checkbox(col_idx, col_letter, col_name, checked=True)
                    found_columns = True
            
            wb.close()
            
            if not found_columns:
                self.add_column_checkbox(6, "G", "(No metadata columns found)", enabled=False)
                self.log.append("⚠️  No non-empty column headers found from column G onwards")
            else:
                self.log.append(f"✓ Loaded {len(self.available_columns)} metadata column(s)")
        
        except Exception as e:
            self.add_column_checkbox(6, "G", f"(Error: {str(e)})", enabled=False)
            self.log.append(f"⚠️  Could not load columns: {e}")
    
    def add_column_checkbox(self, col_idx, col_letter, col_name, checked=False, enabled=True):
        """Add a checkbox for a column."""
        cb = QCheckBox(f"[{col_letter}] {col_name}")
        cb.setChecked(checked)
        cb.setEnabled(enabled)
        cb.setProperty("col_idx", col_idx)
        cb.setProperty("col_name", col_name)
        self.column_checkbox_layout.addWidget(cb)
        self.column_checkboxes.append(cb)
    
    def select_all_columns(self):
        """Check all column checkboxes."""
        for cb in self.column_checkboxes:
            if cb.isEnabled():
                cb.setChecked(True)
    
    def deselect_all_columns(self):
        """Uncheck all column checkboxes."""
        for cb in self.column_checkboxes:
            cb.setChecked(False)
    
    def get_selected_columns(self):
        """Get list of selected (column_index, column_name) tuples."""
        selected = []
        for cb in self.column_checkboxes:
            if cb.isChecked() and cb.isEnabled():
                col_idx = cb.property("col_idx")
                col_name = cb.property("col_name")
                selected.append((col_idx, col_name))
        return selected
    
    def check_resolve_connection(self):
        """Check if DaVinci Resolve API is available."""
        if dvr is None:
            self.log.append("⚠️  DaVinci Resolve API not available")
            self.log.append("    Frame counter feature will not work")
        else:
            try:
                resolve = dvr.scriptapp("Resolve")
                if resolve:
                    project = resolve.GetProjectManager().GetCurrentProject()
                    if project:
                        timeline = project.GetCurrentTimeline()
                        if timeline:
                            self.log.append(f"✓ Connected to Resolve - Timeline: {timeline.GetName()}")
                        else:
                            self.log.append("⚠️  Connected to Resolve but no timeline open")
                    else:
                        self.log.append("⚠️  Connected to Resolve but no project open")
                else:
                    self.log.append("⚠️  Could not connect to Resolve")
            except Exception as e:
                self.log.append(f"⚠️  Resolve connection error: {e}")
    
    def toggle_frame_counter(self, enabled):
        """Enable/disable frame counter inputs."""
        self.fc_file_input.setEnabled(enabled)
        self.browse_fc_btn.setEnabled(enabled)
        self.first_frame.setEnabled(enabled)
    
    def browse_sheet(self):
        """Browse for Excel file."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File",
            self.sheet_input.text(),
            "Excel Files (*.xlsx *.xls)"
        )
        if path:
            self.sheet_input.setText(path)
    
    def browse_output_dir(self):
        """Browse for output directory."""
        directory = QFileDialog.getExistingDirectory(
            self, "Select Output Directory",
            self.output_dir_input.text()
        )
        if directory:
            self.output_dir_input.setText(directory)
    
    def browse_frame_counter(self):
        """Browse for frame counter video file."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Frame Counter Video",
            self.fc_file_input.text(),
            "Video Files (*.mov *.mp4 *.avi *.mxf);;All Files (*.*)"
        )
        if path:
            self.fc_file_input.setText(path)
    
    def start_processing(self):
        """Begin processing metadata and/or frame counters."""
        sheet_path = self.sheet_input.text()
        if not sheet_path or not os.path.exists(sheet_path):
            QMessageBox.warning(self, "Error", "Please specify a valid Excel file")
            return
        
        # Get selected columns
        selected_columns = self.get_selected_columns()
        
        # Validate at least one operation is selected
        fc_enabled = self.fc_enable.isChecked()
        srt_enabled = len(selected_columns) > 0
        
        if not fc_enabled and not srt_enabled:
            QMessageBox.warning(self, "Error", 
                              "Please enable at least one operation:\n"
                              "- Select metadata columns for SRT files, or\n"
                              "- Enable frame counter addition")
            return
        
        # Validate output directory
        output_dir = self.output_dir_input.text()
        if srt_enabled:
            if not output_dir or not os.path.exists(output_dir):
                QMessageBox.warning(self, "Error", "Please specify a valid output directory")
                return
        
        # Validate frame counter settings if enabled
        frame_counter_path = None
        first_frame = None
        if fc_enabled:
            frame_counter_path = self.fc_file_input.text()
            if not frame_counter_path or not os.path.exists(frame_counter_path):
                QMessageBox.warning(self, "Error", "Please specify a valid frame counter video file")
                return
            
            first_frame = self.first_frame.value()
        
        # Get and validate FPS
        fps = self.get_fps()
        if fps is None:
            QMessageBox.warning(self, "Error", "Please enter a valid FPS value (must be positive)")
            return
        
        self.log.clear()
        self.process_btn.setEnabled(False)
        self.progress.setRange(0, 0)
        self.progress.show()
        
        self.worker = MetadataWorker(
            sheet_path, selected_columns, output_dir, fps,
            frame_counter_path, first_frame
        )
        self.worker.progress.connect(self.update_log)
        self.worker.finished.connect(self.processing_done)
        self.worker.start()
    
    def update_log(self, msg):
        """Update log with new message."""
        self.log.append(msg)
        self.log.verticalScrollBar().setValue(
            self.log.verticalScrollBar().maximum()
        )
    
    def processing_done(self, success, msg):
        """Handle processing completion."""
        self.process_btn.setEnabled(True)
        self.progress.hide()
        
        if success:
            QMessageBox.information(self, "Success", msg)
        else:
            QMessageBox.critical(self, "Error", f"Processing failed: {msg}")


def main():
    app = QApplication(sys.argv)
    
    theia_dir = Path("/Library/Application Support/Theia")
    icon_path = theia_dir / "resources" / "graphics" / "add_metadata_icon.png"
    if icon_path.exists():
        icon = QIcon(str(icon_path))
        app.setWindowIcon(icon)
    
    window = AddMetadataGUI()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()