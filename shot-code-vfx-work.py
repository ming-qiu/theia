"""
DaVinci Resolve - Import VFX Shot Codes from Excel to Timeline Subtitles
Reads Excel file and creates SRT subtitle files based on selected columns.
Can also add frame counter videos to timeline shots.
"""

import sys, os
import argparse
from timecode import Timecode
from openpyxl import load_workbook

try:
    import DaVinciResolveScript as dvr
except ImportError:
    dvr = None


def create_srt_file(ws, output_path, column_index, fps):
    """Create SRT file from specific Excel column."""
    
    # Read data from Excel
    subtitles = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= column_index:
            continue
        
        content = row[column_index]
        if content and str(content).strip():
            subtitles.append({
                'tc_in': Timecode(fps, str(row[3])),    # Column D
                'tc_out': Timecode(fps, str(row[4])),   # Column E
                'text': str(content).strip()
            })
    
    if not subtitles:
        return None
    
    # Create SRT content
    srt_lines = []
    for idx, sub in enumerate(subtitles, start=1):
        # Convert to SRT format: HH:MM:SS,mmm
        def to_srt(tc):
            total_sec = tc.frames / tc.framerate
            h = int(total_sec // 3600)
            m = int((total_sec % 3600) // 60)
            s = int(total_sec % 60)
            ms = int((total_sec % 1) * 1000)
            return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"
        
        srt_lines.append(str(idx))
        srt_lines.append(f"{to_srt(sub['tc_in'])} --> {to_srt(sub['tc_out'])}")
        srt_lines.append(sub['text'])
        srt_lines.append("")
    
    # Write SRT file
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(srt_lines))
    
    return output_path


def add_frame_counters(excel_path, frame_counter_path, first_frame, fps):
    """Add frame counter videos to timeline based on shot timings from Excel."""
    
    if not dvr:
        print("ERROR: DaVinci Resolve API not available")
        sys.exit(1)
    
    # Connect to Resolve
    resolve = dvr.scriptapp("Resolve")
    project = resolve.GetProjectManager().GetCurrentProject()
    timeline = project.GetCurrentTimeline()
    mediapool = project.GetMediaPool()
    
    print(f"Timeline: {timeline.GetName()}")
    
    # Import frame counter to media pool
    abs_frame_counter_path = os.path.abspath(frame_counter_path)
    print(f"Importing frame counter: {abs_frame_counter_path}")
    imported = mediapool.ImportMedia([abs_frame_counter_path])
    if not imported:
        print("ERROR: Failed to import frame counter video")
        sys.exit(1)
    
    frame_counter_item = imported[0]
    fc_first_frame = Timecode(fps, frame_counter_item.GetClipProperty("Start TC")).frames - 1

    # Get number of video tracks and create new track
    num_tracks = timeline.GetTrackCount("video")
    target_track = num_tracks + 1
    print(f"Adding frame counters to track {target_track}")
    timeline.AddTrack("video", None)
    
    # Load Excel to get shot timings
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Read shot data
    clips_to_add = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 5:
            continue
        
        record_tc_in = Timecode(fps, str(row[3]))   # Column D
        record_tc_out = Timecode(fps, str(row[4]))  # Column E
        
        shot_duration = record_tc_out.frames - record_tc_in.frames
        
        clips_to_add.append({
            "mediaPoolItem": frame_counter_item,
            "startFrame": first_frame - fc_first_frame,
            "endFrame": first_frame - fc_first_frame + shot_duration,
            "trackIndex": target_track,
            "recordFrame": record_tc_in.frames - 1
        })
    
    print(f"Adding {len(clips_to_add)} frame counter clips...")
    result = mediapool.AppendToTimeline(clips_to_add)
    
    if result:
        print(f"Success! Added {len(result)} frame counter clips to track {target_track}")
    else:
        print("ERROR: Failed to add clips to timeline")


def main():
    parser = argparse.ArgumentParser(description='Create SRT subtitle files from Excel columns')
    parser.add_argument('--excel', default="clip_inventory.xlsx", help='Path to Excel file')
    parser.add_argument('--shot-code', action='store_true', help='Create SRT from column G')
    parser.add_argument('--vfx-work', action='store_true', help='Create SRT from column H')
    parser.add_argument('--vendor', action='store_true', help='Create SRT from column I')
    parser.add_argument('--fps', type=float, default=24.0, help='Timeline FPS (default: 24)')
    parser.add_argument('--frame-counter', type=str, help='Path to frame counter video file')
    parser.add_argument('--first-frame', type=int, help='Starting frame number for frame counters')
    
    args = parser.parse_args()
    
    # Handle frame counter mode
    if args.frame_counter:
        if not args.first_frame:
            print("ERROR: --first-frame required when using --frame-counter")
            sys.exit(1)
        add_frame_counters(args.excel, args.frame_counter, args.first_frame, args.fps)
        return
    
    # Check at least one flag is set for SRT mode
    if not (args.shot_code or args.vfx_work or args.vendor):
        print("ERROR: At least one flag must be specified\n")
        print("Usage examples:")
        print(f"  python {sys.argv[0]} --shot-code")
        print(f"  python {sys.argv[0]} --shot-code --vfx-work --fps 23.976")
        print(f"  python {sys.argv[0]} --frame-counter video.mov --first-frame 1009\n")
        print("Flags:")
        print("  --excel         Path to Excel file (default: clip_inventory.xlsx)")
        print("  --shot-code     Create SRT from column G (VFX Shot Code)")
        print("  --vfx-work      Create SRT from column H (VFX Work)")
        print("  --vendor        Create SRT from column I (Vendor)")
        print("  --fps           Timeline FPS (default: 24)")
        print("  --frame-counter Path to frame counter video")
        print("  --first-frame   Starting frame number (required with --frame-counter)")
        sys.exit(1)
    
    print(f"Loading: {args.excel}")
    print(f"FPS: {args.fps}\n")
    
    # Load Excel once
    wb = load_workbook(args.excel)
    ws = wb.active
    
    # Create SRT files
    base_path = args.excel.replace('.xlsx', '')
    
    if args.shot_code:
        output = f"{base_path}_shot_code.srt"
        if create_srt_file(ws, output, 6, args.fps):
            print(f"Created: {output}")
    
    if args.vfx_work:
        output = f"{base_path}_vfx_work.srt"
        if create_srt_file(ws, output, 7, args.fps):
            print(f"Created: {output}")
    
    if args.vendor:
        output = f"{base_path}_vendor.srt"
        if create_srt_file(ws, output, 8, args.fps):
            print(f"Created: {output}")
    
    print("\nTo import into Resolve:")
    print("1. Go to Edit page")
    print("2. Right-click on subtitle track")
    print("3. Select 'Import Subtitle...'")


if __name__ == "__main__":
    main()