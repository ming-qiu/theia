"""
Theia - Shot Metadata GUI
Import VFX shot codes from Excel to timeline subtitles and add frame counters
"""

import sys
import os
from pathlib import Path
from timecode import Timecode

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QComboBox, QFileDialog,
    QMessageBox, QProgressBar, QTextEdit, QCheckBox, QGroupBox,
    QSpinBox, QDoubleSpinBox, QScrollArea
)
from PySide6.QtCore import Qt, QThread, Signal, QUrl
from PySide6.QtGui import QFont, QDesktopServices, QIcon

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

# Import DaVinci Resolve API
try:
    import DaVinciResolveScript as dvr
except ImportError:
    import sys
    resolve_script_api = "/Library/Application Support/Blackmagic Design/DaVinci Resolve/Developer/Scripting/Modules"
    if resolve_script_api not in sys.path:
        sys.path.append(resolve_script_api)
    try:
        import DaVinciResolveScript as dvr
    except ImportError:
        dvr = None


class MetadataWorker(QThread):
    """Threaded worker for processing metadata and frame counters."""
    progress = Signal(str)
    finished = Signal(bool, str)
    
    def __init__(self, sheet_path, selected_columns, srt_enabled, srt_output_dir, 
                 fcpxml_enabled, fcpxml_output_dir, fps,
                 frame_counter_path=None, first_frame=None):
        super().__init__()
        self.sheet_path = sheet_path
        self.selected_columns = selected_columns  # List of (column_index, column_name) tuples
        self.srt_enabled = srt_enabled
        self.srt_output_dir = srt_output_dir
        self.fcpxml_enabled = fcpxml_enabled
        self.fcpxml_output_dir = fcpxml_output_dir
        self.fps = fps
        self.frame_counter_path = frame_counter_path
        self.first_frame = first_frame
    
    def log(self, msg):
        self.progress.emit(msg)
    
    def create_srt_file(self, ws, output_path, column_index):
        """Create SRT file from specific Excel column."""
        
        # Read data from Excel
        subtitles = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= column_index:
                continue
            
            content = row[column_index]
            if content and str(content).strip():
                subtitles.append({
                    'tc_in': Timecode(self.fps, str(row[3])),    # Column D
                    'tc_out': Timecode(self.fps, str(row[4])),   # Column E
                    'text': str(content).strip()
                })
        
        if not subtitles:
            return None
        
        # Create SRT content
        srt_lines = []
        for idx, sub in enumerate(subtitles, start=1):
            # Convert to SRT format: HH:MM:SS,mmm
            def to_srt(tc):
                total_sec = (tc.frames - 1) / float(tc.framerate)
                h = int(total_sec // 3600)
                m = int((total_sec % 3600) // 60)
                s = int(total_sec % 60)
                ms = int((total_sec % 1) * 1000)
                return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"
            
            srt_lines.append(str(idx))
            srt_lines.append(f"{to_srt(sub['tc_in'])} --> {to_srt(sub['tc_out'])}")
            srt_lines.append(sub['text'])
            srt_lines.append("")
        
        # Write SRT file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(srt_lines))
        
        return output_path
    
    def create_fcpxml_file(self, ws, output_path, column_index, column_name):
        """Create FCPXML file with Basic Title elements from specific Excel column."""
        
        # Read data from Excel
        titles = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= column_index:
                continue
            
            content = row[column_index]
            if content and str(content).strip():
                titles.append({
                    'tc_in': Timecode(self.fps, str(row[3])),    # Column D
                    'tc_out': Timecode(self.fps, str(row[4])),   # Column E
                    'text': str(content).strip()
                })
        
        if not titles:
            return None
        
        # Determine frame rate string for FCPXML
        fps = self.fps
        if abs(fps - 23.976) < 0.001:
            frame_duration = "1001/24000s"
            rate_denominator = 24000
            frame_numerator = 1001
        elif abs(fps - 24) < 0.001:
            frame_duration = "1/24s"
            rate_denominator = 2400
            frame_numerator = 100
        elif abs(fps - 25) < 0.001:
            frame_duration = "1/25s"
            rate_denominator = 2500
            frame_numerator = 100
        elif abs(fps - 29.97) < 0.001:
            frame_duration = "1001/30000s"
            rate_denominator = 30000
            frame_numerator = 1001
        elif abs(fps - 30) < 0.001:
            frame_duration = "1/30s"
            rate_denominator = 3000
            frame_numerator = 100
        elif abs(fps - 60) < 0.001:
            frame_duration = "1/60s"
            rate_denominator = 6000
            frame_numerator = 100
        else:
            # Generic fallback
            frame_duration = f"1/{int(fps)}s"
            rate_denominator = int(fps * 100)
            frame_numerator = 100
        
        # Calculate total duration
        last_out = titles[-1]['tc_out']
        total_duration_frames = last_out.frames * frame_numerator
        
        # Build FCPXML
        lines = []
        lines.append('<?xml version="1.0" encoding="UTF-8"?>')
        lines.append('<!DOCTYPE fcpxml>')
        lines.append('<fcpxml version="1.9">')
        lines.append('  <resources>')
        lines.append(f'    <format id="r1" name="FFVideoFormat1080p{int(fps)}" frameDuration="{frame_duration}" width="1920" height="1080" colorSpace="1-1-1 (Rec. 709)"/>')
        lines.append('    <effect id="r2" name="Basic Title" uid=".../Titles.localized/Bumper:Opener.localized/Basic Title.localized/Basic Title.moti"/>')
        lines.append('  </resources>')
        lines.append('  <library>')
        lines.append(f'    <event name="{column_name}" uid="FA0976C2155BF5E1CD0AA20BD91F88B1">')
        lines.append(f'      <project name="{column_name}" uid="A9CE7D528B481A850DDD48AF2D238B14" modDate="2026-02-02 20:24:21 +0000">')
        
        # Calculate total duration in fractional format
        total_duration_str = f"{total_duration_frames}/{rate_denominator}s"
        lines.append(f'        <sequence format="r1" duration="{total_duration_str}" tcStart="0/{int(fps)}s" tcFormat="NDF" audioLayout="stereo" audioRate="48k">')
        lines.append('          <spine>')
        lines.append(f'            <gap name="Gap" offset="0s" start="0s" duration="{total_duration_str}">')
        
        # Add each title
        for idx, title in enumerate(titles, start=1):
            text = title['text']
            
            # Convert timecodes to frames (0-indexed) and multiply by frame numerator
            offset_frames = (title['tc_in'].frames - 1) * frame_numerator
            start_frames = (title['tc_in'].frames - 1) * frame_numerator
            duration_frames = (title['tc_out'].frames - title['tc_in'].frames) * frame_numerator
            
            # Format as fractions
            offset_str = f"{offset_frames}/{rate_denominator}s"
            start_str = f"{start_frames}/{rate_denominator}s"
            duration_str = f"{duration_frames}/{rate_denominator}s"
            
            ts_id = f"ts{idx}"
            
            lines.append(f'              <title ref="r2" lane="0" name="{text} - Basic Title" offset="{offset_str}" start="{start_str}" duration="{duration_str}">')
            lines.append('                <param name="Flatten" key="9999/999166631/999166633/2/351" value="1"/>')
            lines.append('                <param name="Alignment" key="9999/999166631/999166633/2/354/3142713059/401" value="1 (Center)"/>')
            lines.append('                <param name="Alignment" key="9999/999166631/999166633/2/354/999169573/401" value="1 (Center)"/>')
            lines.append('                <text>')
            lines.append(f'                  <text-style ref="{ts_id}">{text}</text-style>')
            lines.append('                </text>')
            lines.append(f'                <text-style-def id="{ts_id}">')
            lines.append('                  <text-style font="Helvetica" fontSize="60" fontColor="1 1 1 1" alignment="center" fontFace="Regular"/>')
            lines.append('                </text-style-def>')
            lines.append('              </title>')
        
        lines.append('            </gap>')
        lines.append('          </spine>')
        lines.append('        </sequence>')
        lines.append('      </project>')
        lines.append('    </event>')
        lines.append('    <smart-collection name="Projects" match="all">')
        lines.append('      <match-clip rule="is" type="project"/>')
        lines.append('    </smart-collection>')
        lines.append('    <smart-collection name="All Video" match="any">')
        lines.append('      <match-media rule="is" type="videoOnly"/>')
        lines.append('      <match-media rule="is" type="videoWithAudio"/>')
        lines.append('    </smart-collection>')
        lines.append('    <smart-collection name="Audio Only" match="all">')
        lines.append('      <match-media rule="is" type="audioOnly"/>')
        lines.append('    </smart-collection>')
        lines.append('    <smart-collection name="Stills" match="all">')
        lines.append('      <match-media rule="is" type="stills"/>')
        lines.append('    </smart-collection>')
        lines.append('    <smart-collection name="Favorites" match="all">')
        lines.append('      <match-ratings value="favorites"/>')
        lines.append('    </smart-collection>')
        lines.append('  </library>')
        lines.append('</fcpxml>')
        
        # Write FCPXML file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        return output_path
    
    def add_frame_counters(self):
        """Add frame counter videos to timeline based on shot timings from Excel."""
        
        if not dvr:
            self.log("ERROR: DaVinci Resolve API not available")
            return False
        
        try:
            # Connect to Resolve
            resolve = dvr.scriptapp("Resolve")
            project = resolve.GetProjectManager().GetCurrentProject()
            timeline = project.GetCurrentTimeline()
            mediapool = project.GetMediaPool()
            
            self.log(f"Timeline: {timeline.GetName()}")
            
            # Import frame counter to media pool
            abs_frame_counter_path = os.path.abspath(self.frame_counter_path)
            self.log(f"Importing frame counter: {abs_frame_counter_path}")
            imported = mediapool.ImportMedia([abs_frame_counter_path])
            if not imported:
                self.log("ERROR: Failed to import frame counter video")
                return False
            
            frame_counter_item = imported[0]
            fc_first_frame = Timecode(self.fps, frame_counter_item.GetClipProperty("Start TC")).frames - 1
            
            # Get number of video tracks and create new track
            num_tracks = timeline.GetTrackCount("video")
            target_track = num_tracks + 1
            self.log(f"Adding frame counters to track {target_track}")
            timeline.AddTrack("video", None)
            
            # Load Excel to get shot timings
            wb = load_workbook(self.sheet_path)
            ws = wb.active
            
            # Read shot data
            clips_to_add = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) < 5:
                    continue
                
                # Check if any metadata exists in column G onwards (index 6+)
                has_metadata = False
                for cell_value in row[6:]:  # Column G is index 6
                    if cell_value and str(cell_value).strip():
                        has_metadata = True
                        break
                
                # Skip this shot if no metadata
                if not has_metadata:
                    continue
                
                record_tc_in = Timecode(self.fps, str(row[3]))   # Column D
                record_tc_out = Timecode(self.fps, str(row[4]))  # Column E
                
                shot_duration = record_tc_out.frames - record_tc_in.frames
                
                clips_to_add.append({
                    "mediaPoolItem": frame_counter_item,
                    "startFrame": self.first_frame - fc_first_frame,
                    "endFrame": self.first_frame - fc_first_frame + shot_duration,
                    "trackIndex": target_track,
                    "recordFrame": record_tc_in.frames - 1
                })
            
            self.log(f"Adding {len(clips_to_add)} frame counter clips...")
            result = mediapool.AppendToTimeline(clips_to_add)
            
            if result:
                self.log(f"Success! Added {len(result)} frame counter clips to track {target_track}")
                return True
            else:
                self.log("ERROR: Failed to add clips to timeline")
                return False
                
        except Exception as e:
            self.log(f"ERROR: {str(e)}")
            return False
    
    def run(self):
        """Execute the metadata export and/or frame counter addition."""
        try:
            success_count = 0
            
            # Handle frame counter mode
            if self.frame_counter_path and self.first_frame is not None:
                self.log("=" * 50)
                self.log("Adding Frame Counters")
                self.log("=" * 50)
                if self.add_frame_counters():
                    success_count += 1
                else:
                    self.finished.emit(False, "Frame counter addition failed")
                    return
            
            # Create SRT and FCPXML files
            if self.selected_columns and (self.srt_enabled or self.fcpxml_enabled):
                self.log("=" * 50)
                self.log("Creating Subtitle Files")
                self.log("=" * 50)
                
                # Load sheet
                wb = load_workbook(self.sheet_path)
                ws = wb.active
                
                # Create SRT and FCPXML files for each selected column
                srt_count = 0
                fcpxml_count = 0
                for col_idx, col_name in self.selected_columns:
                    # Create SRT file if enabled
                    if self.srt_enabled:
                        srt_output_path = os.path.join(self.srt_output_dir, f"{col_name}.srt")
                        if self.create_srt_file(ws, srt_output_path, col_idx):
                            self.log(f"  Created SRT: {srt_output_path}")
                            srt_count += 1
                    
                    # Create FCPXML file if enabled
                    if self.fcpxml_enabled:
                        fcpxml_output_path = os.path.join(self.fcpxml_output_dir, f"{col_name}.fcpxml")
                        if self.create_fcpxml_file(ws, fcpxml_output_path, col_idx, col_name):
                            self.log(f"  Created FCPXML: {fcpxml_output_path}")
                            fcpxml_count += 1
                
                if srt_count > 0 or fcpxml_count > 0:
                    msg_parts = []
                    if srt_count > 0:
                        msg_parts.append(f"{srt_count} SRT")
                    if fcpxml_count > 0:
                        msg_parts.append(f"{fcpxml_count} FCPXML")
                    self.log(f"Successfully created {' and '.join(msg_parts)} file(s)")
                    success_count += 1
                else:
                    self.log("No files created (no data found in specified columns)")
            
            if success_count > 0:
                self.finished.emit(True, "Processing completed successfully")
            else:
                self.finished.emit(False, "No operations completed")
                
        except Exception as e:
            self.log(f"ERROR: {str(e)}")
            self.finished.emit(False, str(e))


class AddMetadataGUI(QMainWindow):
    """Main GUI window for shot metadata operations."""
    
    def __init__(self):
        super().__init__()
        self.last_export_path = None
        self.setup_ui()

    def setup_ui(self):
        self.setWindowTitle("Theia - Add Metadata")
        self.setMinimumWidth(750)
        self.setMinimumHeight(700)
        
        self.column_checkboxes = []
        self.available_columns = []  # List of (index, name) tuples
        
        # Create central widget
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Title
        title = QLabel("Add Metadata to Timeline")
        font = QFont()
        font.setPointSize(16)
        font.setBold(True)
        title.setFont(font)
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        layout.addSpacing(5)
        
        # Excel file input
        file_group = QGroupBox("Excel File")
        file_layout = QVBoxLayout()
        
        file_row = QHBoxLayout()
        file_row.addWidget(QLabel("Metadata File:"))
        self.sheet_input = QLineEdit(str(Path.home() / "Downloads/clip_inventory.xlsx"))
        self.sheet_input.textChanged.connect(self.on_excel_file_changed)
        file_row.addWidget(self.sheet_input)
        browse_sheet_btn = QPushButton("Browse...")
        browse_sheet_btn.clicked.connect(self.browse_sheet)
        file_row.addWidget(browse_sheet_btn)
        file_layout.addLayout(file_row)
        file_group.setLayout(file_layout)
        layout.addWidget(file_group)
        layout.addSpacing(5)
        
        # Column selection - independent group
        col_group = QGroupBox("Metadata Columns")
        col_layout = QVBoxLayout()
        
        col_header = QHBoxLayout()
        col_header.addWidget(QLabel("Select Columns to Export:"))
        
        select_all_btn = QPushButton("Select All")
        select_all_btn.setMaximumWidth(80)
        select_all_btn.clicked.connect(self.select_all_columns)
        col_header.addWidget(select_all_btn)
        
        deselect_all_btn = QPushButton("Deselect All")
        deselect_all_btn.setMaximumWidth(100)
        deselect_all_btn.clicked.connect(self.deselect_all_columns)
        col_header.addWidget(deselect_all_btn)
        
        refresh_btn = QPushButton("↻")
        refresh_btn.setMaximumWidth(40)
        refresh_btn.setToolTip("Refresh Column List")
        refresh_btn.clicked.connect(self.load_excel_columns)
        col_header.addWidget(refresh_btn)
        
        col_header.addStretch()
        col_layout.addLayout(col_header)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMaximumHeight(150)
        scroll.setMinimumHeight(60)
        
        self.column_checkbox_widget = QWidget()
        self.column_checkbox_layout = QVBoxLayout(self.column_checkbox_widget)
        self.column_checkbox_layout.setContentsMargins(5, 5, 5, 5)
        
        scroll.setWidget(self.column_checkbox_widget)
        col_layout.addWidget(scroll)
        
        col_group.setLayout(col_layout)
        layout.addWidget(col_group)
        layout.addSpacing(5)
        
        # FCPXML export
        fcpxml_group = QGroupBox("FCPXML Titles")
        fcpxml_layout = QVBoxLayout()
        
        self.fcpxml_enable = QCheckBox("Export FCPXML title files")
        self.fcpxml_enable.setChecked(True)
        self.fcpxml_enable.toggled.connect(self.toggle_fcpxml)
        fcpxml_layout.addWidget(self.fcpxml_enable)
        
        fcpxml_output_row = QHBoxLayout()
        fcpxml_output_row.addWidget(QLabel("Output Directory:"))
        self.fcpxml_output_dir_input = QLineEdit(str(Path.home() / "Downloads/"))
        self.fcpxml_output_dir_input.setPlaceholderText("Directory for FCPXML files")
        fcpxml_output_row.addWidget(self.fcpxml_output_dir_input)
        
        browse_fcpxml_output_btn = QPushButton("Browse...")
        browse_fcpxml_output_btn.setMaximumWidth(100)
        browse_fcpxml_output_btn.clicked.connect(self.browse_fcpxml_output_dir)
        self.browse_fcpxml_output_btn = browse_fcpxml_output_btn
        fcpxml_output_row.addWidget(browse_fcpxml_output_btn)
        
        fcpxml_layout.addLayout(fcpxml_output_row)
        fcpxml_group.setLayout(fcpxml_layout)
        layout.addWidget(fcpxml_group)
        layout.addSpacing(5)
        
        # SRT export
        srt_group = QGroupBox("SRT Subtitles")
        srt_layout = QVBoxLayout()
        
        self.srt_enable = QCheckBox("Export SRT subtitle files")
        self.srt_enable.setChecked(True)
        self.srt_enable.toggled.connect(self.toggle_srt)
        srt_layout.addWidget(self.srt_enable)
        
        srt_output_row = QHBoxLayout()
        srt_output_row.addWidget(QLabel("Output Directory:"))
        self.srt_output_dir_input = QLineEdit(str(Path.home() / "Downloads/"))
        self.srt_output_dir_input.setPlaceholderText("Directory for SRT files")
        srt_output_row.addWidget(self.srt_output_dir_input)
        
        browse_srt_output_btn = QPushButton("Browse...")
        browse_srt_output_btn.setMaximumWidth(100)
        browse_srt_output_btn.clicked.connect(self.browse_srt_output_dir)
        self.browse_srt_output_btn = browse_srt_output_btn
        srt_output_row.addWidget(browse_srt_output_btn)
        
        srt_layout.addLayout(srt_output_row)
        srt_group.setLayout(srt_layout)
        layout.addWidget(srt_group)
        layout.addSpacing(5)
        
        # Frame Counter
        fc_group = QGroupBox("Frame Counter")
        fc_layout = QVBoxLayout()
        
        self.fc_enable = QCheckBox("Add frame counter videos to timeline")
        self.fc_enable.toggled.connect(self.toggle_frame_counter)
        fc_layout.addWidget(self.fc_enable)
        
        fc_file_row = QHBoxLayout()
        fc_file_row.addWidget(QLabel("Frame Counter File:"))
        self.fc_file_input = QLineEdit("")
        self.fc_file_input.setPlaceholderText("Path to frame counter video file")
        self.fc_file_input.setEnabled(False)
        fc_file_row.addWidget(self.fc_file_input)
        
        browse_fc_btn = QPushButton("Browse...")
        browse_fc_btn.setMaximumWidth(100)
        browse_fc_btn.clicked.connect(self.browse_frame_counter)
        browse_fc_btn.setEnabled(False)
        self.browse_fc_btn = browse_fc_btn
        fc_file_row.addWidget(browse_fc_btn)
        
        fc_layout.addLayout(fc_file_row)
        
        first_frame_row = QHBoxLayout()
        first_frame_row.addWidget(QLabel("Starting Frame Number:"))
        self.first_frame = QSpinBox()
        self.first_frame.setMinimum(0)
        self.first_frame.setMaximum(999999)
        self.first_frame.setValue(1001)
        self.first_frame.setEnabled(False)
        first_frame_row.addWidget(self.first_frame)
        first_frame_row.addStretch()
        
        fc_layout.addLayout(first_frame_row)
        fc_group.setLayout(fc_layout)
        layout.addWidget(fc_group)
        layout.addSpacing(5)
        
        # Frame Rate Settings
        fps_group = QGroupBox("Set Frame Rate")
        fps_layout = QVBoxLayout()
        
        fps_row = QHBoxLayout()
        fps_row.addWidget(QLabel("Timeline FPS:"))
        
        self.fps_combo = QComboBox()
        self.fps_combo.setMaximumWidth(120)
        self.fps_combo.addItems([
            "23.976",
            "24",
            "25",
            "30",
            "60",
            "Custom..."
        ])
        self.fps_combo.setCurrentText("24")
        self.fps_combo.currentTextChanged.connect(self.on_fps_changed)
        fps_row.addWidget(self.fps_combo)
        
        # Custom FPS input (hidden by default)
        self.custom_fps_input = QLineEdit()
        self.custom_fps_input.setMaximumWidth(100)
        self.custom_fps_input.setPlaceholderText("Enter FPS")
        self.custom_fps_input.setText("")
        self.custom_fps_input.hide()
        self.custom_fps_input.textChanged.connect(self.validate_custom_fps)
        fps_row.addWidget(self.custom_fps_input)
        
        # Show custom input if that was the last selection
        if self.fps_combo.currentText() == "Custom...":
            self.custom_fps_input.show()
        
        fps_row.addStretch()
        
        fps_layout.addLayout(fps_row)
        fps_group.setLayout(fps_layout)
        layout.addWidget(fps_group)
        layout.addSpacing(5)
        
        # Go button
        self.process_btn = QPushButton("Go")
        self.process_btn.setMinimumHeight(40)
        self.process_btn.clicked.connect(self.start_processing)
        layout.addWidget(self.process_btn)
        
        # Progress bar
        self.progress = QProgressBar()
        self.progress.hide()
        layout.addWidget(self.progress)
        
        # Log
        layout.addWidget(QLabel("Log:"))
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        layout.addWidget(self.log)
        
        # Check if Resolve is available
        self.check_resolve_connection()
        
        # Try to load columns from default file
        self.load_excel_columns()
    
    def on_fps_changed(self, text):
        """Handle FPS combo box changes."""
        if text == "Custom...":
            self.custom_fps_input.show()
            self.custom_fps_input.setFocus()
        else:
            self.custom_fps_input.hide()
    
    def validate_custom_fps(self, text):
        """Validate custom FPS input."""
        if not text:
            return
        
        try:
            fps = float(text)
            if fps <= 0:
                self.custom_fps_input.setStyleSheet("background-color: #ffcccc;")
            else:
                self.custom_fps_input.setStyleSheet("")
        except ValueError:
            self.custom_fps_input.setStyleSheet("background-color: #ffcccc;")
    
    def get_fps(self):
        """Get the selected FPS value."""
        if self.fps_combo.currentText() == "Custom...":
            try:
                fps = float(self.custom_fps_input.text())
                if fps <= 0:
                    return None
                return fps
            except ValueError:
                return None
        else:
            return float(self.fps_combo.currentText())
    
    def on_excel_file_changed(self):
        """Handle Excel file path changes."""
        # Auto-load columns when a valid file is entered
        if os.path.exists(self.sheet_input.text()):
            self.load_excel_columns()
    
    def load_excel_columns(self):
        """Load column headers from Excel file starting from column G."""
        # Clear existing checkboxes
        for cb in self.column_checkboxes:
            cb.deleteLater()
        self.column_checkboxes.clear()
        self.available_columns.clear()
        
        sheet_path = self.sheet_input.text()
        if not sheet_path or not os.path.exists(sheet_path):
            self.add_column_checkbox(6, "G", "(No file loaded)", enabled=False)
            return
        
        try:
            wb = load_workbook(sheet_path, read_only=True)
            ws = wb.active
            
            # Read headers from first row, starting at column G (index 7, 0-indexed)
            headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            
            found_columns = False
            for col_idx in range(6, len(headers)):  # Start from column G (index 6)
                header = headers[col_idx]
                if header and str(header).strip():
                    col_letter = get_column_letter(col_idx + 1)
                    col_name = str(header).strip()
                    self.available_columns.append((col_idx, col_name))
                    self.add_column_checkbox(col_idx, col_letter, col_name, checked=True)
                    found_columns = True
            
            wb.close()
            
            if not found_columns:
                self.add_column_checkbox(6, "G", "(No metadata columns found)", enabled=False)
                self.log.append("⚠️  No non-empty column headers found from column G onwards")
            else:
                self.log.append(f"✓ Loaded {len(self.available_columns)} metadata column(s)")
        
        except Exception as e:
            self.add_column_checkbox(6, "G", f"(Error: {str(e)})", enabled=False)
            self.log.append(f"⚠️  Could not load columns: {e}")
    
    def add_column_checkbox(self, col_idx, col_letter, col_name, checked=False, enabled=True):
        """Add a checkbox for a column."""
        cb = QCheckBox(f"[{col_letter}] {col_name}")
        cb.setChecked(checked)
        cb.setEnabled(enabled)
        cb.setProperty("col_idx", col_idx)
        cb.setProperty("col_name", col_name)
        self.column_checkbox_layout.addWidget(cb)
        self.column_checkboxes.append(cb)
    
    def select_all_columns(self):
        """Check all column checkboxes."""
        for cb in self.column_checkboxes:
            if cb.isEnabled():
                cb.setChecked(True)
    
    def deselect_all_columns(self):
        """Uncheck all column checkboxes."""
        for cb in self.column_checkboxes:
            cb.setChecked(False)
    
    def get_selected_columns(self):
        """Get list of selected (column_index, column_name) tuples."""
        selected = []
        for cb in self.column_checkboxes:
            if cb.isChecked() and cb.isEnabled():
                col_idx = cb.property("col_idx")
                col_name = cb.property("col_name")
                selected.append((col_idx, col_name))
        return selected
    
    def check_resolve_connection(self):
        """Check if DaVinci Resolve API is available."""
        if dvr is None:
            self.log.append("⚠️  DaVinci Resolve API not available")
            self.log.append("    Frame counter feature will not work")
        else:
            try:
                resolve = dvr.scriptapp("Resolve")
                if resolve:
                    project = resolve.GetProjectManager().GetCurrentProject()
                    if project:
                        timeline = project.GetCurrentTimeline()
                        if timeline:
                            self.log.append(f"✓ Connected to Resolve - Timeline: {timeline.GetName()}")
                        else:
                            self.log.append("⚠️  Connected to Resolve but no timeline open")
                    else:
                        self.log.append("⚠️  Connected to Resolve but no project open")
                else:
                    self.log.append("⚠️  Could not connect to Resolve")
            except Exception as e:
                self.log.append(f"⚠️  Resolve connection error: {e}")
    
    def toggle_frame_counter(self, enabled):
        """Enable/disable frame counter inputs."""
        self.fc_file_input.setEnabled(enabled)
        self.browse_fc_btn.setEnabled(enabled)
        self.first_frame.setEnabled(enabled)
    
    def toggle_srt(self, enabled):
        """Enable/disable SRT export inputs."""
        self.srt_output_dir_input.setEnabled(enabled)
        self.browse_srt_output_btn.setEnabled(enabled)
    
    def toggle_fcpxml(self, enabled):
        """Enable/disable FCPXML export inputs."""
        self.fcpxml_output_dir_input.setEnabled(enabled)
        self.browse_fcpxml_output_btn.setEnabled(enabled)
    
    def browse_sheet(self):
        """Browse for Excel file."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File",
            self.sheet_input.text(),
            "Excel Files (*.xlsx *.xls)"
        )
        if path:
            self.sheet_input.setText(path)
    
    def browse_srt_output_dir(self):
        """Browse for SRT output directory."""
        directory = QFileDialog.getExistingDirectory(
            self, "Select SRT Output Directory",
            self.srt_output_dir_input.text()
        )
        if directory:
            self.srt_output_dir_input.setText(directory)
    
    def browse_fcpxml_output_dir(self):
        """Browse for FCPXML output directory."""
        directory = QFileDialog.getExistingDirectory(
            self, "Select FCPXML Output Directory",
            self.fcpxml_output_dir_input.text()
        )
        if directory:
            self.fcpxml_output_dir_input.setText(directory)
    
    def browse_frame_counter(self):
        """Browse for frame counter video file."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Frame Counter Video",
            self.fc_file_input.text(),
            "Video Files (*.mov *.mp4 *.avi *.mxf);;All Files (*.*)"
        )
        if path:
            self.fc_file_input.setText(path)
    
    def start_processing(self):
        """Begin processing metadata and/or frame counters."""
        sheet_path = self.sheet_input.text()
        if not sheet_path or not os.path.exists(sheet_path):
            QMessageBox.warning(self, "Error", "Please specify a valid Excel file")
            return
        
        # Get selected columns
        selected_columns = self.get_selected_columns()
        
        # Check which operations are enabled
        fc_enabled = self.fc_enable.isChecked()
        srt_enabled = self.srt_enable.isChecked()
        fcpxml_enabled = self.fcpxml_enable.isChecked()
        
        # Validate at least one operation is selected
        if not fc_enabled and not srt_enabled and not fcpxml_enabled:
            QMessageBox.warning(self, "Error", 
                              "Please enable at least one operation:\n"
                              "- Enable SRT export, or\n"
                              "- Enable FCPXML export, or\n"
                              "- Enable frame counter addition")
            return
        
        # Validate columns selected if SRT or FCPXML is enabled
        if (srt_enabled or fcpxml_enabled) and len(selected_columns) == 0:
            QMessageBox.warning(self, "Error", 
                              "Please select at least one metadata column to export")
            return
        
        # Validate SRT output directory if enabled
        srt_output_dir = None
        if srt_enabled:
            srt_output_dir = self.srt_output_dir_input.text()
            if not srt_output_dir or not os.path.exists(srt_output_dir):
                QMessageBox.warning(self, "Error", "Please specify a valid SRT output directory")
                return
        
        # Validate FCPXML output directory if enabled
        fcpxml_output_dir = None
        if fcpxml_enabled:
            fcpxml_output_dir = self.fcpxml_output_dir_input.text()
            if not fcpxml_output_dir or not os.path.exists(fcpxml_output_dir):
                QMessageBox.warning(self, "Error", "Please specify a valid FCPXML output directory")
                return
        
        # Validate frame counter settings if enabled
        frame_counter_path = None
        first_frame = None
        if fc_enabled:
            frame_counter_path = self.fc_file_input.text()
            if not frame_counter_path or not os.path.exists(frame_counter_path):
                QMessageBox.warning(self, "Error", "Please specify a valid frame counter video file")
                return
            
            first_frame = self.first_frame.value()
        
        # Get and validate FPS
        fps = self.get_fps()
        if fps is None:
            QMessageBox.warning(self, "Error", "Please enter a valid FPS value (must be positive)")
            return
        
        self.log.clear()
        self.process_btn.setEnabled(False)
        self.progress.setRange(0, 0)
        self.progress.show()
        
        self.worker = MetadataWorker(
            sheet_path, selected_columns, srt_enabled, srt_output_dir,
            fcpxml_enabled, fcpxml_output_dir, fps,
            frame_counter_path, first_frame
        )
        self.worker.progress.connect(self.update_log)
        self.worker.finished.connect(self.processing_done)
        self.worker.start()
    
    def update_log(self, msg):
        """Update log with new message."""
        self.log.append(msg)
        self.log.verticalScrollBar().setValue(
            self.log.verticalScrollBar().maximum()
        )
    
    def processing_done(self, success, msg):
        """Handle processing completion."""
        self.process_btn.setEnabled(True)
        self.progress.hide()
        
        if success:
            QMessageBox.information(self, "Success", msg)
        else:
            QMessageBox.critical(self, "Error", f"Processing failed: {msg}")


def main():
    app = QApplication(sys.argv)
    
    theia_dir = Path("/Library/Application Support/Theia")
    icon_path = theia_dir / "resources" / "graphics" / "add_metadata_icon.png"
    if icon_path.exists():
        icon = QIcon(str(icon_path))
        app.setWindowIcon(icon)
    
    window = AddMetadataGUI()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()