"""
Theia - Shot List GUI
Export VFX shot list with elements from a DaVinci Resolve timeline to Excel.
"""

import os
import re
import sys
import traceback
from pathlib import Path
from collections import defaultdict

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QComboBox, QFileDialog,
    QMessageBox, QProgressBar, QTextEdit, QGroupBox, QSpinBox, QScrollArea
)
from PySide6.QtCore import Qt, QThread, Signal, QUrl
from PySide6.QtGui import QFont, QIcon, QDesktopServices

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from timecode import Timecode

# Import DaVinci Resolve API
try:
    import DaVinciResolveScript as dvr
except ImportError:
    resolve_script_api = "/Library/Application Support/Blackmagic Design/DaVinci Resolve/Developer/Scripting/Modules"
    if resolve_script_api not in sys.path:
        sys.path.append(resolve_script_api)
    try:
        import DaVinciResolveScript as dvr
    except ImportError:
        dvr = None

# -------- Utils --------

def get_timeline_fps(timeline):
    fps_str = timeline.GetSetting("timelineFrameRate") or ""
    try:
        fps = float(fps_str)
        return fps
    except Exception:
        return 24.0

def fps_to_str(fps):
    if float(fps).is_integer():
        return str(int(round(fps)))
    return f"{fps:.3f}".rstrip('0').rstrip('.')

def safe_get(d, k, default=None):
    try:
        return d.get(k, default)
    except Exception:
        return default


def element_name_for_track(idx):
    if idx == 1:
        return "ScanBg"
    return f"ScanFg{idx-1:02d}"

def get_sequence_name(shot_code):
    if shot_code.find('_') != -1:
        return shot_code.split('_')[0]
    elif shot_code.find('-') != -1:
        return shot_code.split('-')[0]
    else:
        return 'sequence_name'

def shot_editorial_name_from_bg(bg_elements):
    """Pick the ScanBg reel name for the shot (earliest bg element)."""
    if not bg_elements:
        return ""
    e = sorted(bg_elements, key=lambda x: (x["TimelineStart"], x["TimelineEnd"]))[0]
    return e.get("ReelName", "") or ""

def best_bg_cut_in_tc(bg_elements, fps):
    """Choose Cut In TC from ScanBg (track 1). Prefer the element that starts closest to the shot start."""
    if not bg_elements:
        return ""
    e = sorted(bg_elements, key=lambda x: x["ClipInFrames"])[0]
    return e["ClipInTC"]

def parse_edl(edl_path, fps_str):
    events       = []     # ordered list of event dicts (one per real clip, in EDL order)
    current      = None   # last event (for clip-name attachment)
    m2_target    = None   # event that should receive the next M2 line
    prev_stored  = None   # last event that was appended to events

    with open(edl_path, 'r', encoding='utf-8', errors='replace') as f:
        for raw in f:
            line = raw.rstrip('\n\r')

            # ── C (cut) event line ────────────────────────────────────────────
            # NNN  REEL  TRACK  C  SRC_IN  SRC_OUT  REC_IN  REC_OUT
            m = re.match(
                r'^\d+\s+\S+\s+\S+\s+[Cc]\s+'
                r'(\d{2}:\d{2}:\d{2}[:;]\d{2})\s+'   # src_in
                r'(\d{2}:\d{2}:\d{2}[:;]\d{2})\s+'   # src_out
                r'(\d{2}:\d{2}:\d{2}[:;]\d{2})\s+'   # rec_in
                r'(\d{2}:\d{2}:\d{2}[:;]\d{2})',      # rec_out
                line
            )
            if m:
                src_in, src_out, rec_in, rec_out = [
                    x.replace(';', ':') for x in m.groups()
                ]
                current = {
                    'src_in':     src_in,
                    'src_out':    src_out,
                    'rec_in':     rec_in,
                    'rec_out':    rec_out,
                    'retime_fps': None,
                    'clip_name':  '',
                }
                m2_target = current
                if rec_in != rec_out:
                    # Normal cut — append in EDL order
                    events.append(current)
                    prev_stored = current
                # Zero-duration C lines are dissolve FROM-clip markers.
                # Don't store them (would collide with the D event at the same key),
                # but keep current/m2_target pointing at them so any following M2
                # line correctly targets the FROM clip, not the incoming D clip.
                continue

            # ── D (dissolve) event line ───────────────────────────────────────
            # NNN  REEL  TRACK  D  DUR  SRC_IN  SRC_OUT  REC_IN  REC_OUT
            d = re.match(
                r'^\d+\s+\S+\s+\S+\s+[Dd]\s+(\d+)\s+'
                r'(\d{2}:\d{2}:\d{2}[:;]\d{2})\s+'   # src_in  (incoming clip)
                r'(\d{2}:\d{2}:\d{2}[:;]\d{2})\s+'   # src_out
                r'(\d{2}:\d{2}:\d{2}[:;]\d{2})\s+'   # rec_in  (dissolve start)
                r'(\d{2}:\d{2}:\d{2}[:;]\d{2})',      # rec_out (clip end)
                line
            )
            if d:
                dissolve_len = int(d.group(1))
                src_in, src_out, rec_in, rec_out = [
                    x.replace(';', ':') for x in d.groups()[1:]
                ]
                half = dissolve_len // 2
                # Tag the outgoing clip (last stored event) with its dissolve tail length
                if prev_stored is not None:
                    prev_stored['dissolve_out'] = half
                # Incoming clip event — src_in/rec_in are already at the dissolve start
                current = {
                    'src_in':      src_in,
                    'src_out':     src_out,
                    'rec_in':      rec_in,
                    'rec_out':     rec_out,
                    'retime_fps':  None,
                    'clip_name':   '',
                    'dissolve_in': half,
                }
                events.append(current)
                prev_stored = current
                # m2_target intentionally left pointing at the zero-dur C event
                # (the FROM clip's dissolve-tail retime), not the incoming clip.
                continue

            # ── M2 (motion effect / retime) line ─────────────────────────────
            # M2  REEL  FPS  SRC_TC
            m2 = re.match(r'^M2\s+\S+\s+([\d.]+)', line)
            if m2 and m2_target is not None:
                m2_target['retime_fps'] = float(m2.group(1))
                continue

            # ── FROM CLIP NAME comment ────────────────────────────────────────
            cn = re.match(r'^\*\s*FROM CLIP NAME:\s*(.+)', line)
            if cn and current is not None:
                current['clip_name'] = cn.group(1).strip()
                continue

            # ── TO CLIP NAME comment (dissolve incoming clip) ─────────────────
            to_cn = re.match(r'^\*\s*TO CLIP NAME:\s*(.+)', line)
            if to_cn and current is not None:
                current['clip_name'] = to_cn.group(1).strip()

    return events


def get_clip_tc_from_edl(timeline_item, fps, edl_event=None):
    fps_str = fps_to_str(fps)
    try:
        src_in_frames  = Timecode(fps_str, edl_event['src_in']).frames - 1   # 0-indexed inclusive
        src_out_frames = Timecode(fps_str, edl_event['src_out']).frames - 2  # EDL out is exclusive
        rec_in_frames  = Timecode(fps_str, edl_event['rec_in']).frames - 1
        rec_out_frames = Timecode(fps_str, edl_event['rec_out']).frames - 1
        dur = rec_out_frames - rec_in_frames

        dissolve_out = edl_event.get('dissolve_out', 0)
        if dissolve_out:
            src_out_frames += dissolve_out * 2
            rec_out_frames += dissolve_out * 2
            dur            += dissolve_out * 2
        return {
            "ClipInTC":      repr(Timecode(fps_str, frames=max(1, src_in_frames + 1))),
            "ClipInFrames":  src_in_frames,
            "ClipOutTC":     repr(Timecode(fps_str, frames=max(1, src_out_frames + 1))),
            "ClipOutFrames": src_out_frames,
            "ClipDuration":  dur,
        }
    except Exception:
        return read_clip_tc(timeline_item, fps)

def read_clip_tc(timeline_item, fps):

    fps_str = fps_to_str(fps)
    dur     = int(timeline_item.GetDuration())
    
    # ── MediaPoolItem path ─────────────────────────────────────────────────
    mpi = timeline_item.GetMediaPoolItem()
    if mpi:
        try:
            props        = mpi.GetClipProperty() or {}
            src_fps_str  = props.get("FPS") or fps_str
            src_fps      = float(src_fps_str)
            start_tc_str = props.get("Start TC") or "00:00:00:00"
            # 0-indexed absolute frame of the clip's first source frame
            mpi_start    = Timecode(fps_to_str(src_fps), start_tc_str).frames - 1
            src_in_frames  = mpi_start + int(timeline_item.GetSourceStartFrame())
            src_out_frames = src_in_frames + dur - 1
            return {
                "ClipInTC":      repr(Timecode(fps_str, frames=max(1, src_in_frames + 1))),
                "ClipInFrames":  src_in_frames,
                "ClipOutTC":     repr(Timecode(fps_str, frames=max(1, src_out_frames + 1))),
                "ClipOutFrames": src_out_frames,
                "ClipDuration":  dur,
            }
        except Exception:
            pass  # fall through

    # ── Float-time fallback ────────────────────────────────────────────────
    tin = timeline_item.GetSourceStartTime()
    if tin is not None:
        src_in_frames  = int(round(tin * fps))
    else:
        src_in_frames  = 0
    src_out_frames = src_in_frames + dur - 1

    return {
        "ClipInTC":      repr(Timecode(fps_str, frames=max(1, src_in_frames + 1))),
        "ClipInFrames":  src_in_frames,
        "ClipOutTC":     repr(Timecode(fps_str, frames=max(1, src_out_frames + 1))),
        "ClipOutFrames": src_out_frames,
        "ClipDuration":  dur,
    }


def _is_back_to_back(prev_src_out, curr_src_in, tol=1):
    return curr_src_in == prev_src_out or curr_src_in == prev_src_out + 1 or abs(curr_src_in - prev_src_out) <= tol

def _fmt_percent(val):
    f = float(str(val).strip())
    p = f * 100.0
    return f"{int(round(p))}%" if abs(p - round(p)) < 1e-6 else f"{p:.2f}%"

def retime_summary(elements_by_track, fps, scan_handle):
    for track_num, track in elements_by_track.items():
        track.sort(key=lambda e: (e["TimelineStart"], e["TimelineEnd"]))

        for clip in track:
            edl_event = clip.get("EDLEvent")

            retime_fps = edl_event.get('retime_fps')
            if retime_fps is not None:
                # CMX 3600: M2 fps = source frames per timeline second
                speed     = retime_fps / fps
                has_retime = abs(speed - 1.0) > 1e-3
            else:
                # No M2 line → definitively 100% speed
                speed      = 1.0
                has_retime = False

        merged_track = []
        i = 0
        while i < len(track):
            group = [track[i]]
            reel = track[i].get("ReelName", "")
            clip_duration = track[i].get("ClipDuration", 0)
            j = i + 1

            while j < len(track):
                same_reel = (track[j].get("ReelName", "") == reel)
                if not same_reel:
                    break
                if not _is_back_to_back(group[-1]["ClipOutFrames"], track[j]["ClipInFrames"]):
                    break
                group.append(track[j])
                clip_duration += track[j].get("ClipDuration", 0)
                j += 1

            any_retime = any(g["HasRetime"] for g in group)

            if any_retime:
                first = group[0]
                last = group[-1]

                parts = []
                total_length = 0
                for k in range(0, len(group)):
                    seg_retime_fps = group[k].get("RetimeFPS") or fps
                    n = round(group[k].get("ClipDuration") * seg_retime_fps / fps)
                    parts.append(f"{n} @ {fps_to_str(seg_retime_fps)}")
                    total_length += n
                
                updated_clip_out_frames = first["ClipInFrames"] + total_length - 1
                updated_clip_out_tc = repr(Timecode(fps_to_str(fps), frames = (updated_clip_out_frames + 1)))
                updated_tail_out = first["ClipIn"] + total_length + scan_handle - 1

                if len(group) > 1:
                    summary = ", ".join(parts)
                    merged_element = {
                        "TrackIndex": first["TrackIndex"],
                        "ShotCode": first["ShotCode"],
                        "ElementName": first["ElementName"],
                        "TimelineItem": first["TimelineItem"],
                        "TimelineStart": first["TimelineStart"],
                        "TimelineEnd": last["TimelineEnd"],
                        "ClipIn": first["ClipIn"],
                        "ClipOut": last["ClipOut"],
                        "ClipInTC": first["ClipInTC"],
                        "ClipOutTC": updated_clip_out_tc,
                        "ClipInFrames": first["ClipInFrames"],
                        "ClipOutFrames": updated_clip_out_frames,
                        "ClipDuration": total_length,
                        "RetimeSummary": summary,
                        "ScaleRepo": first["ScaleRepo"],
                        "HasRetime": True,
                        "ReelName": reel,
                        "Props": first["Props"],
                        "HeadIn": first["HeadIn"],
                        "TailOut": updated_tail_out,
                        "Speed": None,
                        "RetimeFPS": None,
                        "EDLEvent": first.get("EDLEvent"),
                    }
                    merged_track.append(merged_element)
                else:
                    for seg in group:
                        if seg["HasRetime"]:
                            seg["RetimeSummary"] = _fmt_percent(seg['Speed'])
                            seg["ClipOutFrames"] = updated_clip_out_frames
                            seg["ClipOutTC"] = updated_clip_out_tc
                            seg["ClipDuration"] = total_length
                            seg["TailOut"] = updated_tail_out
                        merged_track.append(seg)
            else:
                merged_track.extend(group)

            i = j

        elements_by_track[track_num] = merged_track

def summarize_scale_repo(props):
    zx = safe_get(props, "ZoomX")
    px = safe_get(props, "Pan")
    py = safe_get(props, "Tilt")
    parts = []
    if zx is not None and str(zx).strip() not in ('', 'None') and zx != 1:
        parts.append(f"Scale: {_fmt_percent(zx)}")
    if (px is not None or py is not None) and (abs(px) > 1e-2 or abs(py) > 1e-2):
        parts.append(f"Repo: {px:.2f},{py:.2f}")
    return " ".join(parts)

# -------- Old Excel loading --------

def load_old_shot_list_excel(excel_path, fps):
    """
    Read the Shots sheet from an old shot list Excel.
    Returns dict mapping ShotCode -> {CutIn, CutOut, CutInTC, CutInTCFrames, ...}

    Columns: A=Sequence, B=Cut Order, C=Editorial Name, D=Shot Code,
    E=Change to Cut, F=Work In, G=Cut In, H=Cut Out, I=Work Out,
    J=Cut Duration, K=Bg Retime, L=Fg Retime, M=Cut In TC
    """
    wb = load_workbook(excel_path, read_only=True)
    try:
        ws = wb["Shots"]
    except KeyError:
        ws = wb.active

    old_shots = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 13:
            continue
        shot_code = row[3]  # Column D
        if not shot_code or not str(shot_code).strip():
            continue

        cut_in = int(row[6]) if row[6] is not None else None
        cut_out = int(row[7]) if row[7] is not None else None
        cut_in_tc = str(row[12]).strip() if row[12] else ""

        cut_in_tc_frames = None
        if cut_in_tc:
            try:
                cut_in_tc_frames = Timecode(fps_to_str(fps), cut_in_tc).frames
            except Exception:
                pass

        old_shots[str(shot_code).strip()] = {
            'CutIn': cut_in,
            'CutOut': cut_out,
            'CutInTC': cut_in_tc,
            'CutInTCFrames': cut_in_tc_frames,
        }

    wb.close()
    return old_shots

def compare_with_old_excel(current_shots, old_shots_dict):
    """Compare current shots against old Excel data. Modifies current_shots in place."""
    for shot in current_shots:
        code = shot['ShotCode']
        old = old_shots_dict.get(code)
        if not old or old['CutIn'] is None or old['CutOut'] is None:
            continue

        d_in = shot['CutIn'] - old['CutIn']
        d_out = shot['CutOut'] - old['CutOut']
        change_to_cut = ""
        if d_in != 0 or d_out != 0:
            change_to_cut = (
                f"In: {d_in}, Out: {d_out}" if (d_in != 0 and d_out != 0)
                else f"In: {d_in}" if d_in != 0
                else f"Out: {d_out}"
            )
        shot['ChangeToCut'] = change_to_cut

# -------- Worker --------

class ShotListWorker(QThread):
    """Threaded worker for shot list extraction and Excel export."""
    progress = Signal(str)
    finished = Signal(bool, str)

    def __init__(self, frame_counter_track, track_edl_map,
                 old_excel_path, work_handle, scan_handle,
                 output_path, input_sequence, fps):
        super().__init__()
        self.frame_counter_track = frame_counter_track
        self.track_edl_map = track_edl_map  # dict: track_num -> edl_path
        self.old_excel_path = old_excel_path
        self.work_handle = work_handle
        self.scan_handle = scan_handle
        self.output_path = output_path
        self.input_sequence = input_sequence
        self.fps = fps

    def log(self, msg):
        self.progress.emit(msg)

    def run(self):
        try:
            self.log("Connecting to DaVinci Resolve...")
            if dvr is None:
                self.finished.emit(False, "DaVinci Resolve API not available")
                return

            resolve = dvr.scriptapp("Resolve")
            if not resolve:
                self.finished.emit(False, "Could not connect to DaVinci Resolve")
                return

            project = resolve.GetProjectManager().GetCurrentProject()
            if not project:
                self.finished.emit(False, "No project open in Resolve")
                return

            timeline = project.GetCurrentTimeline()
            if not timeline:
                self.finished.emit(False, "No timeline open in Resolve")
                return

            fps = self.fps
            fps_str = fps_to_str(fps)
            self.log(f"Timeline: {timeline.GetName()} | FPS: {fps}")

            # Load user-specified EDLs per video track
            edl_by_track = {}
            for track_num, edl_path in self.track_edl_map.items():
                if edl_path and os.path.exists(edl_path):
                    self.log(f"Loading EDL for track {track_num}: {os.path.basename(edl_path)}")
                    edl_by_track[track_num] = parse_edl(edl_path, fps_str)
                else:
                    edl_by_track[track_num] = {}

            # Load old shot list Excel if provided
            old_shots_dict = None
            if self.old_excel_path:
                self.log(f"Loading old shot list: {os.path.basename(self.old_excel_path)}")
                old_shots_dict = load_old_shot_list_excel(self.old_excel_path, fps)
                self.log(f"  Found {len(old_shots_dict)} shots in old Excel")

            # Get frame counter clips (define shot boundaries, shot codes, and frame numbers)
            fc_items = timeline.GetItemListInTrack("video", self.frame_counter_track) or []
            if not fc_items:
                self.finished.emit(False,
                    f"No clips found on Frame Counter Track {self.frame_counter_track}")
                return
            fc_items_sorted = sorted(fc_items, key=lambda c: c.GetStart(False))
            self.log(f"Found {len(fc_items_sorted)} frame counter clips on track {self.frame_counter_track}")

            def lookup_edl_event(track_num, index):
                event_list = edl_by_track.get(track_num, [])
                if index < 0 or index >= len(event_list):
                    return None
                return event_list[index]

            # Pre-pull video items for element tracks
            element_labels = {i: element_name_for_track(i) for i in self.track_edl_map}

            skip_tracks = {self.frame_counter_track}
            element_tracks = sorted(k for k in self.track_edl_map if k not in skip_tracks)

            track_items = {}
            # Resolve transition names — these appear as timeline items but are not clips
            # and have no entry in the EDL. Extend this set if new transition types appear.
            RESOLVE_TRANSITIONS = {
                "Cross Dissolve", "Dip to Color Dissolve", "Film Dissolve",
                "Smooth Cut", "Dip to Black", "Dip to White",
            }

            for i in element_tracks:
                all_items = timeline.GetItemListInTrack("video", i) or []
                # Exclude built-in transitions; keep all other items (clips, generators,
                # placeholders) so index alignment with the EDL is preserved.
                track_items[i] = [it for it in all_items if it.GetName() not in RESOLVE_TRANSITIONS]

            # Process each shot
            shots_rows = []
            elements_rows = []
            cut_order = 0

            # {something}_in / out: VFX frame number, out is inclusive
            # {something}_start / end: raw frame number, end is non-inclusive
            for fc_item in fc_items_sorted:
                shot_code = (fc_item.GetName() or "").strip()
                if not shot_code:
                    continue

                cut_order += 1
                shot_start = fc_item.GetStart(False)
                shot_end = fc_item.GetEnd(False)
                shot_dur = shot_end - shot_start

                self.log(f"==== Cut {cut_order}: {shot_code} [{shot_start}-{shot_end}] ====")
                fc_tc_info = read_clip_tc(fc_item, fps)
                cut_in = fc_tc_info['ClipInFrames']
                cut_out = cut_in + int(shot_dur) - 1
                self.log(f"  Cut In={cut_in}, Cut Out={cut_out}")

                # Collect elements on [bottom..top] tracks
                elements_by_track = defaultdict(list)
                first_elem_in_shot = True

                for track in element_tracks:

                    for edl_idx, elem in enumerate(track_items.get(track) or []):
                        elem_start = elem.GetStart(False)
                        elem_end   = elem.GetEnd(False)

                        # Skip disabled elems
                        try:
                            if not elem.GetClipEnabled():
                                continue
                        except Exception:
                            pass

                        mpi = elem.GetMediaPoolItem()
                        mpi_props = (mpi.GetClipProperty() if mpi else {}) or {}
                        reel = mpi_props["File Name"] if mpi_props != {} else elem.GetName()
                        elem_edl_event = lookup_edl_event(track, edl_idx)

                        if elem_edl_event is None:
                            continue

                        # Prefer the EDL clip_name — it's always the actual source file name.
                        if elem_edl_event.get('clip_name'):
                            reel = elem_edl_event['clip_name']

                        if elem_start >= shot_start and elem_end <= shot_end:

                            tc_info = get_clip_tc_from_edl(elem, fps, elem_edl_event)

                            elem_dur = tc_info["ClipDuration"]
                            elem_in = int(fc_tc_info['ClipInFrames'] + elem_start - shot_start)

                            if "dissolve_in" in elem_edl_event:
                                if first_elem_in_shot:
                                    elem_in -= elem_edl_event['dissolve_in']
                                cut_in -= elem_edl_event['dissolve_in']
                            elem_out = int(elem_in + elem_dur - 1)

                            if "dissolve_out" in elem_edl_event:
                                cut_out += elem_edl_event['dissolve_out']
                            first_elem_in_shot = False

                            retime_fps = elem_edl_event.get('retime_fps')
                            if retime_fps is not None:
                                # CMX 3600: M2 fps = source frames per timeline second
                                speed     = retime_fps / fps
                                has_retime = abs(speed - 1.0) > 1e-3
                            else:
                                # No M2 line → definitively 100% speed
                                speed      = 1.0
                                has_retime = False

                            props = elem.GetProperty() or {}

                            elements_by_track[track].append({
                                "TrackIndex":    track,
                                "ShotCode":      shot_code,
                                "ElementName":   element_labels[track],
                                "TimelineItem":  elem,
                                "TimelineStart": elem_start,
                                "TimelineEnd":   elem_end,
                                "ClipIn":        elem_in,
                                "ClipOut":       elem_out,
                                "ClipInTC":      tc_info["ClipInTC"],
                                "ClipOutTC":     tc_info["ClipOutTC"],
                                "ClipInFrames":  tc_info["ClipInFrames"],
                                "ClipOutFrames": tc_info["ClipOutFrames"],
                                "ClipDuration":  elem_dur,
                                "HasRetime":     has_retime,
                                "Speed":         speed,
                                "RetimeSummary": "",
                                "RetimeFPS":     (fps * speed) if speed is not None else None,
                                "ScaleRepo":     summarize_scale_repo(props),
                                "ReelName":      reel,
                                "Props":         props,
                                "HeadIn":        int(elem_in  - self.scan_handle),
                                "TailOut":       int(elem_out + self.scan_handle),
                                "EDLEvent":      elem_edl_event,
                            })

                # Shot metadata from BG elements (lowest element track)
                bottom_track = element_tracks[0] if element_tracks else None
                bg_elems = elements_by_track.get(bottom_track, [])
                cut_in_tc = best_bg_cut_in_tc(bg_elems, fps) if bg_elems else ""
                shot_editorial_name = shot_editorial_name_from_bg(bg_elems)

                work_in = int(cut_in - self.work_handle)
                work_out = int(cut_out + self.work_handle)

                retime_summary(elements_by_track, fps, self.scan_handle)

                bg_retime = "x" if any(e["HasRetime"] for e in elements_by_track.get(bottom_track, [])) else ""
                fg_retime = "x" if any(
                    e["HasRetime"]
                    for track, lst in elements_by_track.items()
                    if track != bottom_track for e in lst
                ) else ""

                # Determine sequence name by input and shot code
                if self.input_sequence == None:
                    sequence = get_sequence_name(shot_code)
                else:
                    sequence = self.input_sequence

                shots_rows.append({
                    "Sequence": sequence,
                    "CutOrder": cut_order,
                    "EditorialName": shot_editorial_name,
                    "ShotCode": shot_code,
                    "ChangeToCut": None,
                    "WorkIn": work_in,
                    "CutIn": int(cut_in),
                    "CutOut": int(cut_out),
                    "WorkOut": work_out,
                    "CutDuration": int(cut_out - cut_in + 1),
                    "BgRetime": bg_retime,
                    "FgRetime": fg_retime,
                    "CutInTC": cut_in_tc,
                })

                for track in element_tracks:
                    for e in sorted(elements_by_track.get(track, []), key=lambda x: (x["TimelineStart"], x["TimelineEnd"])):
                        elements_rows.append({
                            "Sequence": sequence,
                            "CutOrder": cut_order,
                            "EditorialName": e["ReelName"],
                            "ShotCode": shot_code,
                            "Element": e["ElementName"],
                            "ShotCutIn": int(cut_in),
                            "ShotCutOut": int(cut_out),
                            "ClipInTC": e["ClipInTC"],
                            "ClipInFrames": int(e["ClipInFrames"]),
                            "ClipIn": int(e["ClipIn"]),
                            "ClipOut": int(e["ClipOut"]),
                            "ClipOutFrames": int(e["ClipOutFrames"]),
                            "ClipOutTC": e["ClipOutTC"],
                            "ClipHeadIn": int(e["HeadIn"]),
                            "ClipTailOut": int(e["TailOut"]),
                            "ClipDuration": e['ClipDuration'],
                            "Retime": e["RetimeSummary"],
                            "ScaleRepo": e["ScaleRepo"],
                        })

            # Sort
            shots_rows.sort(key=lambda r: r["CutOrder"])
            elements_rows.sort(key=lambda r: r["CutOrder"])

            # Compare with old Excel
            if old_shots_dict:
                compare_with_old_excel(shots_rows, old_shots_dict)

            # -------- Excel output --------
            self.log("")
            self.log("Writing Excel...")

            wb = Workbook()
            ws_shots = wb.active
            ws_shots.title = "Shots"
            shots_cols = [
                "Sequence", "Cut Order", "Editorial Name", "Shot Code", "Change to Cut",
                "Work In", "Cut In", "Cut Out", "Work Out",
                "Cut Duration", "Bg Retime", "Fg Retime", "Cut In TC"
            ]
            ws_shots.append(shots_cols)
            for r in shots_rows:
                ws_shots.append([
                    r["Sequence"], r["CutOrder"], r["EditorialName"], r["ShotCode"], r["ChangeToCut"],
                    r["WorkIn"], r["CutIn"], r["CutOut"], r["WorkOut"],
                    r["CutDuration"], r["BgRetime"], r["FgRetime"], r["CutInTC"]
                ])

            ws_elems = wb.create_sheet(title="Elements")
            elems_cols = [
                "Sequence", "Cut Order", "Editorial Name", "Shot Code", "Element",
                "Cut In", "Cut Out", "Clip Duration (with dissolve before retime)",
                "Clip In TC", "Clip In Frames", "Clip In", "Clip Out", "Clip Out Frames", "Clip Out TC",
                "ScanIn", "ScanOut", "Retime Summary", "Scale & Repo"
            ]
            ws_elems.append(elems_cols)
            for r in elements_rows:
                ws_elems.append([
                    r["Sequence"], r["CutOrder"], r["EditorialName"], r["ShotCode"], r["Element"],
                    r["ShotCutIn"], r["ShotCutOut"], r['ClipDuration'],
                    r["ClipInTC"], r["ClipInFrames"], r["ClipIn"], r["ClipOut"], r["ClipOutFrames"], r["ClipOutTC"],
                    r["ClipHeadIn"], r["ClipTailOut"], r["Retime"], r["ScaleRepo"]
                ])

            # Auto-width
            for ws in (ws_shots, ws_elems):
                for col_idx, _ in enumerate(ws[1], start=1):
                    max_len = 0
                    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                        val = row[0].value
                        if val is None:
                            continue
                        max_len = max(max_len, len(str(val)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 50)

            out_path = os.path.abspath(self.output_path)
            wb.save(out_path)
            self.log(f"✓ Wrote Excel: {out_path}")
            self.log(f"  {len(shots_rows)} shots, {len(elements_rows)} elements")

            self.finished.emit(True, f"Exported {len(shots_rows)} shots to {os.path.basename(out_path)}")

        except Exception as e:
            self.log(f"ERROR: {e}")
            self.log(traceback.format_exc())
            self.finished.emit(False, str(e))

# -------- GUI --------

class ShotListGUI(QMainWindow):
    """Main GUI window for shot list export."""

    def __init__(self):
        super().__init__()
        self.worker = None
        self.setup_ui()
        self.populate_tracks()

    def setup_ui(self):
        self.setWindowTitle("Theia - Shot List")
        self.setMinimumWidth(750)
        self.setMinimumHeight(700)

        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # Title
        title = QLabel("Shot List")
        font = QFont()
        font.setPointSize(16)
        font.setBold(True)
        title.setFont(font)
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        layout.addSpacing(5)

        # ---- Track Configuration ----
        track_group = QGroupBox("Track Configuration")
        track_layout = QVBoxLayout()

        # Frame Counter Track
        fc_row = QHBoxLayout()
        fc_row.addWidget(QLabel("Frame Counter Track:"))
        self.counter_track_combo = QComboBox()
        self.counter_track_combo.setMinimumWidth(200)
        fc_row.addWidget(self.counter_track_combo)
        fc_row.addStretch()
        track_layout.addLayout(fc_row)

        # Track EDL assignments
        edl_header_row = QHBoxLayout()
        edl_header_row.addWidget(QLabel("Assign an EDL file for each track to process:"))
        refresh_btn = QPushButton("↻")
        refresh_btn.setMaximumWidth(40)
        refresh_btn.setToolTip("Refresh track list from Resolve")
        refresh_btn.clicked.connect(self.populate_tracks)
        edl_header_row.addWidget(refresh_btn)
        edl_header_row.addStretch()
        track_layout.addLayout(edl_header_row)

        self.track_rows = {}  # track_num -> QLineEdit
        self.track_list_widget = QWidget()
        self.track_list_layout = QVBoxLayout(self.track_list_widget)
        self.track_list_layout.setSpacing(4)
        self.track_list_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidget(self.track_list_widget)
        scroll.setWidgetResizable(True)
        scroll.setMaximumHeight(400)
        track_layout.addWidget(scroll)
        track_group.setLayout(track_layout)
        layout.addWidget(track_group)

        # ---- Sequence Name Setting ----
        seq_group = QGroupBox("Sequence Name")
        seq_layout = QVBoxLayout()

        seq_row = QHBoxLayout()
        seq_row.addWidget(QLabel("Sequence Name:"))
        self.seq_name = QLineEdit("")
        self.seq_name.setPlaceholderText("Optional")
        seq_row.addWidget(self.seq_name)
        seq_row.addStretch()
        seq_layout.addLayout(seq_row)
        seq_group.setLayout(seq_layout)

        layout.addWidget(seq_group)

        # ---- Frame Number Settings ----
        frame_group = QGroupBox("Frame Number Settings")
        frame_layout = QVBoxLayout()

        old_row = QHBoxLayout()
        old_row.addWidget(QLabel("Old Shot List:"))
        self.old_excel_input = QLineEdit("")
        self.old_excel_input.setPlaceholderText("Optional - previous shot list .xlsx for comparison & frame mapping")
        old_row.addWidget(self.old_excel_input)
        browse_old_btn = QPushButton("Browse...")
        browse_old_btn.setMaximumWidth(100)
        browse_old_btn.clicked.connect(self.browse_old_excel)
        old_row.addWidget(browse_old_btn)
        frame_layout.addLayout(old_row)

        frame_group.setLayout(frame_layout)
        layout.addWidget(frame_group)

        # ---- Handles ----
        handle_group = QGroupBox("Handles")
        handle_layout = QHBoxLayout()

        handle_layout.addWidget(QLabel("Work Handle:"))
        self.work_handle_spin = QSpinBox()
        self.work_handle_spin.setMinimum(0)
        self.work_handle_spin.setMaximum(999)
        self.work_handle_spin.setValue(8)
        handle_layout.addWidget(self.work_handle_spin)

        handle_layout.addSpacing(20)

        handle_layout.addWidget(QLabel("Scan Handle:"))
        self.scan_handle_spin = QSpinBox()
        self.scan_handle_spin.setMinimum(0)
        self.scan_handle_spin.setMaximum(999)
        self.scan_handle_spin.setValue(24)
        handle_layout.addWidget(self.scan_handle_spin)

        handle_layout.addStretch()
        handle_group.setLayout(handle_layout)
        layout.addWidget(handle_group)

        # ---- Output ----
        output_group = QGroupBox("Output")
        output_layout = QHBoxLayout()

        output_layout.addWidget(QLabel("Output File:"))
        self.output_input = QLineEdit(str(Path.home() / "Downloads" / "shot_list.xlsx"))
        output_layout.addWidget(self.output_input)
        browse_output_btn = QPushButton("Browse...")
        browse_output_btn.setMaximumWidth(100)
        browse_output_btn.clicked.connect(self.browse_output)
        output_layout.addWidget(browse_output_btn)

        output_group.setLayout(output_layout)
        layout.addWidget(output_group)

        # ---- Frame Rate ----
        fps_group = QGroupBox("Frame Rate")
        fps_layout = QHBoxLayout()

        fps_layout.addWidget(QLabel("Timeline FPS:"))
        self.fps_combo = QComboBox()
        self.fps_combo.setMaximumWidth(120)
        self.fps_combo.addItems(["23.976", "24", "25", "30", "60", "Custom..."])
        self.fps_combo.setCurrentText("24")
        self.fps_combo.currentTextChanged.connect(self.on_fps_changed)
        fps_layout.addWidget(self.fps_combo)

        self.custom_fps_input = QLineEdit()
        self.custom_fps_input.setMaximumWidth(100)
        self.custom_fps_input.setPlaceholderText("Enter FPS")
        self.custom_fps_input.hide()
        self.custom_fps_input.textChanged.connect(self.validate_custom_fps)
        fps_layout.addWidget(self.custom_fps_input)

        fps_layout.addStretch()
        fps_group.setLayout(fps_layout)
        layout.addWidget(fps_group)

        # ---- Go button ----
        self.go_btn = QPushButton("Go")
        self.go_btn.setMinimumHeight(40)
        self.go_btn.clicked.connect(self.start_processing)
        layout.addWidget(self.go_btn)

        # Progress bar
        self.progress = QProgressBar()
        self.progress.hide()
        layout.addWidget(self.progress)

        # Log
        layout.addWidget(QLabel("Log:"))
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        layout.addWidget(self.log)

    def populate_tracks(self):
        """Populate track dropdowns from the current Resolve timeline."""
        self.counter_track_combo.clear()

        if dvr is None:
            self.log.append("⚠️  DaVinci Resolve API not available")
            return

        try:
            self.log.clear()
            resolve = dvr.scriptapp("Resolve")
            if not resolve:
                self.log.append("⚠️  Could not connect to Resolve")
                return

            project = resolve.GetProjectManager().GetCurrentProject()
            if not project:
                self.log.append("⚠️  No project open")
                return

            timeline = project.GetCurrentTimeline()
            if not timeline:
                self.log.append("⚠️  No timeline open")
                return

            track_count = timeline.GetTrackCount("video")
            self.log.append(f"✓ Timeline: {timeline.GetName()} ({track_count} video tracks)")

            for i in range(track_count, 0, -1):
                name = timeline.GetTrackName("video", i) or ""
                label = f"Track {i}" + (f" ({name})" if name else "")
                self.counter_track_combo.addItem(label, i)

            # Rebuild track EDL rows (preserve existing paths)
            old_paths = {num: inp.text() for num, inp in self.track_rows.items()}
            self.track_rows.clear()
            while self.track_list_layout.count():
                item = self.track_list_layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()

            for i in range(track_count, 0, -1):
                name = timeline.GetTrackName("video", i) or ""
                label_text = f"Track {i}" + (f"  ({name})" if name else "")

                row_widget = QWidget()
                row = QHBoxLayout(row_widget)
                row.setContentsMargins(0, 2, 0, 2)
                lbl = QLabel(label_text)
                lbl.setMinimumWidth(140)
                row.addWidget(lbl)
                edl_input = QLineEdit(old_paths.get(i, ""))
                edl_input.setPlaceholderText("No EDL")
                row.addWidget(edl_input)
                browse_btn = QPushButton("Browse EDL...")
                browse_btn.setMaximumWidth(110)
                browse_btn.clicked.connect(lambda checked, inp=edl_input: self._browse_edl(inp))
                row.addWidget(browse_btn)

                self.track_rows[i] = edl_input
                self.track_list_layout.addWidget(row_widget)

            # Try to auto-detect FPS from timeline
            try:
                fps = get_timeline_fps(timeline)
                fps_str = fps_to_str(fps)
                idx = self.fps_combo.findText(fps_str)
                if idx >= 0:
                    self.fps_combo.setCurrentIndex(idx)
                else:
                    self.fps_combo.setCurrentText("Custom...")
                    self.custom_fps_input.setText(str(fps))
                    self.custom_fps_input.show()
            except Exception:
                pass

        except Exception as e:
            self.log.append(f"⚠️  Resolve connection error: {e}")

    def _browse_edl(self, line_edit):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select EDL",
            line_edit.text() or str(Path.home() / "Downloads"),
            "EDL Files (*.edl);;All Files (*)"
        )
        if path:
            line_edit.setText(path)

    def on_fps_changed(self, text):
        if text == "Custom...":
            self.custom_fps_input.show()
            self.custom_fps_input.setFocus()
        else:
            self.custom_fps_input.hide()

    def validate_custom_fps(self, text):
        if not text:
            return
        try:
            fps = float(text)
            if fps <= 0:
                self.custom_fps_input.setStyleSheet("background-color: #ffcccc;")
            else:
                self.custom_fps_input.setStyleSheet("")
        except ValueError:
            self.custom_fps_input.setStyleSheet("background-color: #ffcccc;")

    def get_fps(self):
        if self.fps_combo.currentText() == "Custom...":
            try:
                fps = float(self.custom_fps_input.text())
                if fps <= 0:
                    return None
                return fps
            except ValueError:
                return None
        else:
            return float(self.fps_combo.currentText())

    def browse_old_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Old Shot List",
            self.old_excel_input.text() or str(Path.home() / "Downloads"),
            "Excel Files (*.xlsx *.xls)"
        )
        if path:
            self.old_excel_input.setText(path)

    def browse_output(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Shot List As",
            self.output_input.text(),
            "Excel Files (*.xlsx)"
        )
        if path:
            self.output_input.setText(path)

    def start_processing(self):
        fps = self.get_fps()
        if fps is None:
            QMessageBox.warning(self, "Error", "Please enter a valid FPS value")
            return

        frame_counter_track = self.counter_track_combo.currentData()
        if frame_counter_track is None:
            QMessageBox.warning(self, "Error", "Please select a Frame Counter Track")
            return

        output = self.output_input.text()
        if not output:
            QMessageBox.warning(self, "Error", "Please specify an output file")
            return

        old_excel = self.old_excel_input.text().strip()
        if old_excel and not os.path.exists(old_excel):
            QMessageBox.warning(self, "Error", "Old Shot List file not found")
            return
        old_excel_path = old_excel if old_excel else None

        self.log.clear()
        self.go_btn.setEnabled(False)
        self.progress.setRange(0, 0)
        self.progress.show()
        self.last_export_path = output

        track_edl_map = {
            num: inp.text().strip()
            for num, inp in self.track_rows.items()
            if num != frame_counter_track and inp.text().strip()
        }

        self.worker = ShotListWorker(
            frame_counter_track=frame_counter_track,
            track_edl_map=track_edl_map,
            old_excel_path=old_excel_path,
            work_handle=self.work_handle_spin.value(),
            scan_handle=self.scan_handle_spin.value(),
            output_path=output,
            input_sequence=self.seq_name.text() or None,
            fps=fps,
        )
        self.worker.progress.connect(self.update_log)
        self.worker.finished.connect(self.processing_done)
        self.worker.start()

    def update_log(self, msg):
        self.log.append(msg)
        self.log.verticalScrollBar().setValue(
            self.log.verticalScrollBar().maximum()
        )

    def processing_done(self, success, msg):
        self.go_btn.setEnabled(True)
        self.progress.hide()

        if success:
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Success")
            msg_box.setText(msg)
            msg_box.setIcon(QMessageBox.Information)
            open_btn = msg_box.addButton("Open File", QMessageBox.ActionRole)
            msg_box.addButton(QMessageBox.Ok)
            msg_box.exec()
            if msg_box.clickedButton() == open_btn:
                if self.last_export_path and Path(self.last_export_path).exists():
                    QDesktopServices.openUrl(QUrl.fromLocalFile(self.last_export_path))
        else:
            QMessageBox.critical(self, "Error", f"Export failed: {msg}")


def main():
    app = QApplication(sys.argv)

    theia_dir = Path("/Library/Application Support/Theia")
    icon_path = theia_dir / "resources" / "graphics" / "shot_list_icon.png"
    if icon_path.exists():
        app.setWindowIcon(QIcon(str(icon_path)))

    window = ShotListGUI()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
