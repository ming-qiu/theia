"""
Theia - Clip Inventory GUI
Exports DaVinci Resolve timeline clips to Excel with thumbnails
"""
import sys
import base64
import time
from io import BytesIO
from pathlib import Path
from timecode import Timecode

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QComboBox, QFileDialog,
    QMessageBox, QProgressBar, QTextEdit, QCheckBox, QScrollArea
)
from PySide6.QtCore import Qt, QThread, Signal, QUrl
from PySide6.QtGui import QFont, QDesktopServices, QIcon

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

# Import DaVinci Resolve API
try:
    import DaVinciResolveScript as dvr
except ImportError:
    import sys
    resolve_script_api = "/Library/Application Support/Blackmagic Design/DaVinci Resolve/Developer/Scripting/Modules"
    if resolve_script_api not in sys.path:
        sys.path.append(resolve_script_api)
    try:
        import DaVinciResolveScript as dvr
    except ImportError:
        dvr = None


class IntervalSet:
    """Helper class for efficient interval operations."""
    def __init__(self, intervals):
        self.intervals = sorted(intervals)
    
    def intersect(self, start, end):
        """Return portions of (start, end) that overlap with this set."""
        result = []
        for (s, e) in self.intervals:
            overlap_start = max(start, s)
            overlap_end = min(end, e)
            if overlap_start < overlap_end:
                result.append((overlap_start, overlap_end))
        return result
    
    def subtract(self, start, end):
        """Remove (start, end) from this interval set."""
        new_intervals = []
        for (s, e) in self.intervals:
            if end <= s or start >= e:
                new_intervals.append((s, e))
            elif start > s and end < e:
                new_intervals.append((s, start))
                new_intervals.append((end, e))
            elif start > s:
                new_intervals.append((s, start))
            elif end < e:
                new_intervals.append((end, e))
        self.intervals = new_intervals
    
    def is_empty(self):
        return len(self.intervals) == 0


class ExportWorker(QThread):
    """Threaded export worker to keep GUI responsive."""
    progress = Signal(str)
    finished = Signal(bool, str, int)
    
    def __init__(self, output_path, selected_tracks):
        super().__init__()
        self.output_path = output_path
        self.selected_tracks = selected_tracks  # List of track numbers
    
    def log(self, msg):
        self.progress.emit(msg)
    
    def get_timeline_duration_frames(self, timeline):
        """Get timeline duration in frames."""
        return timeline.GetEndFrame() - timeline.GetStartFrame()

    def get_visible_clips(self, timeline):
        """Get all visible clips considering track layering, but ignore transitions.
        Handles dissolves/transitions by classifying them into:
          - type1: connecting two clips on same track (treat as hard cut)
          - type2: sitting at end of a clip (ignore transition)
          - type3: sitting at beginning of a clip (apply 0.5 * dissolve length to that clip's start)
        Returns a list of dicts: { 'clip', 'track_num', 'clip_start', 'clip_end', 'visible_ranges' }
        """
        track_count = timeline.GetTrackCount("video")

        # Determine track range
        if self.selected_tracks:
            min_track = min(self.selected_tracks)
            max_track = max(self.selected_tracks)
        else:
            min_track, max_track = 1, track_count

        timeline_start = timeline.GetStartFrame()
        timeline_end = timeline.GetEndFrame()

        self.log(f"Timeline range: {timeline_start} to {timeline_end}")

        # Helper: detect transition items (best-effort)
        def is_transition_item(tl_item):
            try:
                name = (tl_item.GetName() or "").lower()
                if any(k in name for k in ("dissolve", "transition", "wipe", "fade")):
                    return True
                return False
            except Exception:
                return False

        # Pre-scan to compute per-clip start adjustments for type-3 (beginning-of-clip) transitions
        start_adjustments = {}    # clip_id -> frames to add to clip start
        transition_ids = set()    # skip these items later

        for track_num in range(1, track_count + 1):
            # skip tracks not in selection if the user selected tracks
            if self.selected_tracks and track_num not in self.selected_tracks:
                continue

            clips = timeline.GetItemListInTrack("video", track_num) or []
            # iterate with index so we can inspect neighbours
            for i, item in enumerate(clips):
                if not is_transition_item(item):
                    continue

                transition_ids.add(id(item))
                trans_start = item.GetStart()
                trans_end = item.GetEnd()
                trans_len = max(0, int(trans_end - trans_start))

                prev_item = clips[i - 1] if i - 1 >= 0 else None
                next_item = clips[i + 1] if i + 1 < len(clips) else None

                # check simple overlaps with prev/next to classify
                overlaps_prev = prev_item is not None and (prev_item.GetEnd() > trans_start)
                overlaps_next = next_item is not None and (next_item.GetStart() < trans_end)

                # Type1: overlaps both -> connecting two clips on same track (ignore for now)
                if overlaps_prev and overlaps_next:
                    self.log(f"  Transition on track {track_num} connecting two clips ({trans_start}-{trans_end}) -> type1 (ignore)")
                    continue

                # Type2: overlaps previous only -> sitting at end of a clip (do nothing)
                if overlaps_prev and not overlaps_next:
                    self.log(f"  Transition on track {track_num} at end of clip ({trans_start}-{trans_end}) -> type2 (ignore)")
                    continue

                # Type3: overlaps next only -> transition at beginning of clip; nudge next clip start
                if overlaps_next and not overlaps_prev and next_item is not None:
                    adj = int(round(0.5 * trans_len))
                    next_id = id(next_item)
                    # ensure we do not push start past end
                    clip_start = next_item.GetStart()
                    clip_end = next_item.GetEnd()
                    max_allowed = max(0, clip_end - clip_start - 1)
                    adj = min(adj, max_allowed)
                    if adj > 0:
                        start_adjustments[next_id] = start_adjustments.get(next_id, 0) + adj
                        self.log(f"  Transition on track {track_num} at start of next clip ({trans_start}-{trans_end}) -> type3: adding {adj} frames to clip id {next_id}")
                    else:
                        self.log(f"  Transition on track {track_num} at start of next clip ({trans_start}-{trans_end}) -> type3 but adj=0 (too short)")

                    continue

                # If neither overlaps (weird case), just ignore
                self.log(f"  Transition on track {track_num} ({trans_start}-{trans_end}) unclassified -> ignored")

        # Now run occlusion pass top-down, but skip transitions and apply start_adjustments
        visible_regions = IntervalSet([(timeline_start, timeline_end)])
        visible_clips = []

        for track_num in range(max_track, min_track - 1, -1):
            # Skip if track not selected
            if self.selected_tracks and track_num not in self.selected_tracks:
                self.log(f"  Skipping unselected track {track_num}")
                continue

            clips = timeline.GetItemListInTrack("video", track_num)
            if not clips:
                self.log(f"  Track {track_num}: No clips")
                continue

            self.log(f"  Track {track_num}: Processing {len(clips)} clips")

            for clip in clips:
                cid = id(clip)
                # skip transitions entirely
                if cid in transition_ids or is_transition_item(clip):
                    self.log(f"    Skipping transition item {clip.GetName()} [{clip.GetStart()}-{clip.GetEnd()}]")
                    continue

                raw_start = clip.GetStart()
                raw_end = clip.GetEnd()

                # Apply any start adjustment (from type3 transitions)
                adj = start_adjustments.get(cid, 0)
                eff_start = raw_start + adj
                # clamp to not exceed end
                if eff_start >= raw_end:
                    self.log(f"    Clip {clip.GetName()} adjusted start >= end after adj ({raw_start}+{adj} >= {raw_end}) -> skipping")
                    continue

                # Get visible portions of this clip (using adjusted start)
                clip_visible = visible_regions.intersect(eff_start, raw_end)

                if clip_visible:
                    self.log(f"    {clip.GetName()} [{raw_start}-{raw_end}] (effective start {eff_start}): VISIBLE {clip_visible}")
                    visible_clips.append({
                        'clip': clip,
                        'track_num': track_num,
                        'clip_start': eff_start,
                        'clip_end': raw_end,
                        'visible_ranges': clip_visible
                    })

                    # Remove this clip's effective area from visible regions
                    visible_regions.subtract(eff_start, raw_end)
                else:
                    self.log(f"    {clip.GetName()} [{raw_start}-{raw_end}] (effective start {eff_start}): OCCLUDED")

            if visible_regions.is_empty():
                self.log(f"  All regions occluded, stopping at track {track_num}")
                break

        # Sort by timeline position (earliest first)
        visible_clips.sort(key=lambda x: x['clip_start'])

        return visible_clips

    
    def get_thumbnail(self, timeline, frame, fps, target_track_num=None):
        """Get thumbnail at a specific timeline frame.

        NOTE: Resolve's GetCurrentClipThumbnailImage() is tied to the *current* video item.
        When multiple items overlap at the playhead, Resolve may pick an unexpected item.
        To make the selection deterministic, we temporarily solo the requested video track.
        """
        track_states = None
        try:
            # Convert timeline frame -> timeline timecode (Timecode lib is 1-based frames)
            tc = Timecode(fps, frames=int(frame + 1))

            # If requested, temporarily solo the target track so the "current video item" is unambiguous
            if target_track_num is not None:
                track_count = timeline.GetTrackCount("video")
                track_states = {}
                for t in range(1, track_count + 1):
                    # Cache current enable state
                    try:
                        track_states[t] = bool(timeline.GetIsTrackEnabled("video", t))
                    except Exception:
                        # If the call fails for some reason, assume enabled and continue
                        track_states[t] = True

                # Solo the requested track (disable all others)
                for t in range(1, track_count + 1):
                    timeline.SetTrackEnable("video", t, (t == int(target_track_num)))

            # Move playhead
            if not timeline.SetCurrentTimecode(str(tc)):
                return None

            # Give UI a tiny moment to update in Color page (helps on heavy timelines)
            time.sleep(0.03)

            # Fetch thumbnail for Resolve's current video item
            thumb = timeline.GetCurrentClipThumbnailImage()
            if not thumb or not thumb.get('data'):
                return None

            img_bytes = base64.b64decode(thumb['data'])
            img = PILImage.frombytes('RGB', (thumb['width'], thumb['height']), img_bytes)

            # Resize for Excel
            aspect = img.width / img.height if img.height else 1.0
            new_h = 150
            new_w = max(1, int(new_h * aspect))
            return img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)

        except Exception as e:
            self.log(f"    Thumbnail error: {e}")
            return None

        finally:
            # Restore original track enable states
            if track_states:
                for t, enabled in track_states.items():
                    try:
                        timeline.SetTrackEnable("video", t, enabled)
                    except Exception:
                        pass



    def run(self):
        """Main export logic."""
        try:
            self.log("Connecting to DaVinci Resolve...")
            
            if dvr is None:
                self.finished.emit(False, "DaVinci Resolve API not available. Make sure Resolve is running.", 0)
                return
            
            resolve = dvr.scriptapp("Resolve")
            if not resolve:
                self.finished.emit(False, "Could not connect to Resolve. Make sure DaVinci Resolve is running.", 0)
                return
            
            project = resolve.GetProjectManager().GetCurrentProject()
            timeline = project.GetCurrentTimeline() if project else None
            
            if not timeline:
                self.finished.emit(False, "No timeline open", 0)
                return
            
            # Save current state
            starting_page = resolve.GetCurrentPage()
            starting_timecode = timeline.GetCurrentTimecode()
            self.log(f"Saving playhead position: {starting_timecode}")
            # Save video track enable states (we may temporarily solo tracks for thumbnails)
            starting_video_track_states = {}
            try:
                video_track_count = timeline.GetTrackCount("video")
                for t in range(1, video_track_count + 1):
                    starting_video_track_states[t] = bool(timeline.GetIsTrackEnabled("video", t))
            except Exception:
                starting_video_track_states = {}

            
            # Switch to Color page for thumbnails
            self.log("Opening Color page...")
            resolve.OpenPage('color')
            time.sleep(0.5)
            
            # Get timeline info
            timeline_name = timeline.GetName()
            fps = float(timeline.GetSetting("timelineFrameRate"))
            self.log(f"Timeline: {timeline_name} ({fps} fps)")
            
            # Get visible clips using occlusion logic
            track_str = ", ".join(str(t) for t in sorted(self.selected_tracks))
            self.log(f"Analyzing tracks: {track_str}")
            
            visible_clips = self.get_visible_clips(timeline)
            
            if not visible_clips:
                # Restore video track enable states
                if starting_video_track_states:
                    for t, enabled in starting_video_track_states.items():
                        try:
                            timeline.SetTrackEnable("video", t, enabled)
                        except Exception:
                            pass
                resolve.OpenPage(starting_page)
                timeline.SetCurrentTimecode(starting_timecode)
                self.finished.emit(False, f"No visible clips found on selected tracks", 0)
                return
            
            self.log(f"Found {len(visible_clips)} visible clips\n")
            
            # Flatten visible clips into individual visible ranges and sort by start time
            all_visible_ranges = []
            for clip_info in visible_clips:
                clip = clip_info['clip']
                track_num = clip_info['track_num']
                visible_ranges = clip_info['visible_ranges']
                
                for vis_start, vis_end in visible_ranges:
                    all_visible_ranges.append({
                        'clip': clip,
                        'track_num': track_num,
                        'vis_start': vis_start,
                        'vis_end': vis_end,
                        'clip_start_record': clip_info['clip_start'],
                        'clip_start_api': clip_info.get('api_start', clip_info['clip_start'])
                    })
            
            # Sort by visible range start time
            all_visible_ranges.sort(key=lambda x: x['vis_start'])
            
            self.log(f"Expanded to {len(all_visible_ranges)} visible range(s)\n")
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Shots"
            
            # Headers
            headers = ["Thumbnail", "Reel Name", "Cut Order", "Record In", 
                      "Record Out", "Source In", "Notes"]
            for idx, header in enumerate(headers, 1):
                ws.cell(1, idx, header)
            
            # Column widths
            ws.column_dimensions['A'].width = 34
            ws.column_dimensions['B'].width = 25
            for col in ['C', 'D', 'E', 'F', 'G']:
                ws.column_dimensions[col].width = 15
            
            # Process clips - one row per visible range
            row_num = 2  # Start after header
            
            # Track which clips have multiple visible ranges for naming
            clip_range_counts = {}
            clip_range_indices = {}
            
            for range_info in all_visible_ranges:
                clip = range_info['clip']
                clip_id = id(clip)  # Use object id as unique identifier
                
                if clip_id not in clip_range_counts:
                    # Count how many ranges this clip has
                    clip_range_counts[clip_id] = sum(1 for r in all_visible_ranges if id(r['clip']) == clip_id)
                    clip_range_indices[clip_id] = 0
                
                clip_range_indices[clip_id] += 1
            
            # Reset indices for actual processing
            clip_range_indices = {}
            
            for cut_order, range_info in enumerate(all_visible_ranges, 1):
                clip = range_info['clip']
                track_num = range_info['track_num']
                vis_start = range_info['vis_start']
                vis_end = range_info['vis_end']
                clip_start_record = range_info['clip_start_record']
                clip_start_api = range_info['clip_start_api']
                clip_id = id(clip)
                
                # Track which part this is
                if clip_id not in clip_range_indices:
                    clip_range_indices[clip_id] = 0
                clip_range_indices[clip_id] += 1
                
                self.log(f"[{cut_order}] Track {track_num}: {clip.GetName()} [{vis_start}-{vis_end}]")
                
                # Reel Name
                clip_name = clip.GetName()
                if clip_range_counts[clip_id] > 1:
                    clip_name = f"{clip_name} (part {clip_range_indices[clip_id]})"
                ws.cell(row_num, 2, clip_name)
                
                # Cut Order
                ws.cell(row_num, 3, cut_order)
                
                # Timecodes - use visible range
                tc_in = Timecode(fps, frames=int(vis_start + 1))
                tc_out = Timecode(fps, frames=int(vis_end + 1))
                ws.cell(row_num, 4, str(tc_in))
                ws.cell(row_num, 5, str(tc_out))
                
                # Source timecode
                media_item = clip.GetMediaPoolItem()
                if media_item:
                    try:
                        props = media_item.GetClipProperty()
                        src_fps = float(props.get('FPS', fps))
                        src_tc_str = props.get('Start TC')
                        src_start = Timecode(str(src_fps), src_tc_str).frames
                        
                        # Calculate source TC for this visible range start
                        offset_into_clip = vis_start - clip_start_record
                        src_tc = Timecode(src_fps, frames=int(src_start + clip.GetLeftOffset() + offset_into_clip))
                        ws.cell(row_num, 6, str(src_tc))
                    except:
                        ws.cell(row_num, 6, str(tc_in))
                else:
                    ws.cell(row_num, 6, str(tc_in))
                
                # Thumbnail - grab at the visible start frame
                if media_item:
                    thumb_frame = max(int(vis_start), int(clip_start_api))
                    thumb = self.get_thumbnail(timeline, thumb_frame, fps, target_track_num=track_num)
                    if thumb:
                        buf = BytesIO()
                        thumb.save(buf, format='PNG')
                        buf.seek(0)
                        ws.add_image(XLImage(buf), f'A{row_num}')
                        ws.row_dimensions[row_num].height = 112.5
                    else:
                        ws.cell(row_num, 1, "No thumbnail")
                else:
                    ws.cell(row_num, 1, "No thumbnail")
                
                row_num += 1
            
            total_rows = len(all_visible_ranges)
            
            # Save
            self.log(f"\nSaving to {self.output_path}...")
            wb.save(self.output_path)
            
            # Restore video track enable states
            if starting_video_track_states:
                for t, enabled in starting_video_track_states.items():
                    try:
                        timeline.SetTrackEnable("video", t, enabled)
                    except Exception:
                        pass

            # Restore timeline state
            self.log(f"Restoring playhead to: {starting_timecode}")
            timeline.SetCurrentTimecode(starting_timecode)
            resolve.OpenPage(starting_page)
            
            self.finished.emit(True, f"Successfully exported {len(all_visible_ranges)} clips", len(all_visible_ranges))
            
        except Exception as e:
            self.log(f"\nERROR: {e}")
            import traceback
            self.log(traceback.format_exc())
            # Best-effort restore of playhead/page/track states
            try:
                if 'starting_video_track_states' in locals() and starting_video_track_states:
                    for t, enabled in starting_video_track_states.items():
                        try:
                            timeline.SetTrackEnable("video", t, enabled)
                        except Exception:
                            pass
                if 'starting_timecode' in locals():
                    timeline.SetCurrentTimecode(starting_timecode)
                if 'starting_page' in locals():
                    resolve.OpenPage(starting_page)
            except Exception:
                pass

            self.finished.emit(False, str(e), 0)


class ClipInventoryGUI(QMainWindow):
    """Main GUI window for Clip Inventory."""
    
    def __init__(self):
        super().__init__()
        self.worker = None
        self.last_export_path = None
        self.track_checkboxes = []
        self.setup_ui()
        self.populate_track_list()
    
    def setup_ui(self):
        """Build the interface."""
        self.setWindowTitle("Theia - Clip Inventory")
        self.setMinimumSize(600, 600)
        
        # Main layout
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Title
        title = QLabel("Export Timeline to Excel")
        font = QFont()
        font.setPointSize(16)
        font.setBold(True)
        title.setFont(font)
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Track selection section
        track_section = QWidget()
        track_layout = QVBoxLayout(track_section)
        track_layout.setContentsMargins(0, 0, 0, 0)
        
        # Track selection header
        track_header = QHBoxLayout()
        track_header.addWidget(QLabel("Video Tracks:"))
        
        select_all_btn = QPushButton("Select All")
        select_all_btn.setMaximumWidth(80)
        select_all_btn.clicked.connect(self.select_all_tracks)
        track_header.addWidget(select_all_btn)
        
        deselect_all_btn = QPushButton("Deselect All")
        deselect_all_btn.setMaximumWidth(90)
        deselect_all_btn.clicked.connect(self.deselect_all_tracks)
        track_header.addWidget(deselect_all_btn)
        
        refresh_btn = QPushButton("↻")
        refresh_btn.setMaximumWidth(40)
        refresh_btn.setToolTip("Refresh track list")
        refresh_btn.clicked.connect(self.populate_track_list)
        track_header.addWidget(refresh_btn)
        
        track_header.addStretch()
        track_layout.addLayout(track_header)
        
        # Scrollable checkbox area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMaximumHeight(150)
        scroll.setMinimumHeight(60)
        
        self.track_checkbox_widget = QWidget()
        self.track_checkbox_layout = QVBoxLayout(self.track_checkbox_widget)
        self.track_checkbox_layout.setContentsMargins(5, 5, 5, 5)
        
        scroll.setWidget(self.track_checkbox_widget)
        track_layout.addWidget(scroll)
        
        layout.addWidget(track_section)
        
        # Output file
        file_row = QHBoxLayout()
        file_row.addWidget(QLabel("Output File:"))
        self.file_input = QLineEdit(str(Path.home() / "Downloads/clip_inventory.xlsx"))
        file_row.addWidget(self.file_input)
        browse = QPushButton("Browse...")
        browse.clicked.connect(self.browse_file)
        file_row.addWidget(browse)
        layout.addLayout(file_row)
        
        # Export button
        self.export_btn = QPushButton("Export")
        self.export_btn.setMinimumHeight(40)
        self.export_btn.clicked.connect(self.start_export)
        layout.addWidget(self.export_btn)
        
        # Progress bar
        self.progress = QProgressBar()
        self.progress.hide()
        layout.addWidget(self.progress)
        
        # Log
        layout.addWidget(QLabel("Log:"))
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        layout.addWidget(self.log)
    
    def populate_track_list(self):
        """Get available video tracks from current timeline."""
        # Clear existing checkboxes
        for cb in self.track_checkboxes:
            cb.deleteLater()
        self.track_checkboxes.clear()
        
        try:
            if dvr is None:
                self.add_track_checkbox(1, checked=True)
                return
            
            resolve = dvr.scriptapp("Resolve")
            if not resolve:
                self.add_track_checkbox(1, checked=True)
                return
            
            project = resolve.GetProjectManager().GetCurrentProject()
            timeline = project.GetCurrentTimeline() if project else None
            
            if not timeline:
                self.add_track_checkbox(1, checked=True)
                self.log.append("⚠️  No timeline open - defaulting to track 1")
                return
            
            # Get track count
            track_count = timeline.GetTrackCount("video")
            
            if track_count == 0:
                self.add_track_checkbox(1, checked=True)
                self.log.append("⚠️  No video tracks found - defaulting to track 1")
                return
            
            # Create checkboxes for each track
            for i in range(1, track_count + 1):
                self.add_track_checkbox(i, checked=True)
            
            self.log.append(f"✓ Found {track_count} video track(s)")
            
        except Exception as e:
            self.add_track_checkbox(1, checked=True)
            self.log.append(f"⚠️  Could not get tracks: {e}")
    
    def add_track_checkbox(self, track_num, checked=True):
        """Add a checkbox for a track."""
        cb = QCheckBox(f"Track {track_num}")
        cb.setChecked(checked)
        cb.setProperty("track_num", track_num)
        self.track_checkbox_layout.addWidget(cb)
        self.track_checkboxes.append(cb)
    
    def select_all_tracks(self):
        """Check all track checkboxes."""
        for cb in self.track_checkboxes:
            cb.setChecked(True)
    
    def deselect_all_tracks(self):
        """Uncheck all track checkboxes."""
        for cb in self.track_checkboxes:
            cb.setChecked(False)
    
    def get_selected_tracks(self):
        """Get list of selected track numbers."""
        selected = []
        for cb in self.track_checkboxes:
            if cb.isChecked():
                selected.append(cb.property("track_num"))
        return selected
    
    def browse_file(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Inventory", 
            self.file_input.text(),
            "Excel Files (*.xlsx)"
        )
        if path:
            self.file_input.setText(path)
    
    def start_export(self):
        """Begin export process."""
        output = self.file_input.text()
        if not output:
            QMessageBox.warning(self, "Error", "Please specify output file")
            return
        
        # Get selected tracks
        selected_tracks = self.get_selected_tracks()
        if not selected_tracks:
            QMessageBox.warning(self, "Error", "Please select at least one track")
            return
        
        self.log.clear()
        self.export_btn.setEnabled(False)
        self.progress.setRange(0, 0)
        self.progress.show()
        
        # Store path for later
        self.last_export_path = output
        
        self.worker = ExportWorker(output, selected_tracks)
        self.worker.progress.connect(self.update_log)
        self.worker.finished.connect(self.export_done)
        self.worker.start()
    
    def update_log(self, msg):
        self.log.append(msg)
        self.log.verticalScrollBar().setValue(
            self.log.verticalScrollBar().maximum()
        )
    
    def export_done(self, success, msg, clip_count):
        """Handle export completion."""
        self.export_btn.setEnabled(True)
        self.progress.hide()
        
        if success:
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Export Complete")
            msg_box.setText(f"Successfully exported {clip_count} clip(s)")
            msg_box.setIcon(QMessageBox.Information)
            
            open_btn = msg_box.addButton("Open File", QMessageBox.ActionRole)
            close_btn = msg_box.addButton(QMessageBox.Ok)
            
            msg_box.exec()
            
            if msg_box.clickedButton() == open_btn:
                self.open_export_file()
        else:
            QMessageBox.critical(self, "Error", f"Export failed: {msg}")
    
    def open_export_file(self):
        """Open the exported Excel file in default application."""
        if self.last_export_path and Path(self.last_export_path).exists():
            QDesktopServices.openUrl(QUrl.fromLocalFile(self.last_export_path))
        else:
            QMessageBox.warning(self, "Error", "Export file not found")


def main():
    app = QApplication(sys.argv)
    
    theia_dir = Path("/Library/Application Support/Theia")
    icon_path = theia_dir / "resources" / "graphics" / "clip_inventory_icon.png"
    if icon_path.exists():
        icon = QIcon(str(icon_path))
        app.setWindowIcon(icon)
    
    window = ClipInventoryGUI()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()